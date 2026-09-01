"""Verified evidence-only exports for archived Agent runs.

Reports are projections of the bounded manifest stored by ``AnalysisHistoryStore``.
They never reopen the source dataset, execute a tool or include unverified model prose.
"""

from __future__ import annotations

import hashlib
import html
import json
import math
import re
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any, Literal

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, PageBreak, SimpleDocTemplate, Spacer, Table, TableStyle

from lacopilot.analyst_document_reports import (
    _pdf_paragraph,
    _pdf_styles,
    _pdf_table_style,
    _register_fonts,
)
from lacopilot.audit import audit
from lacopilot.config import get_settings
from lacopilot.tools.common import safe_excel_writer, safe_output_path

AgentReportFormat = Literal["xlsx", "html", "pdf"]

XLSX_REPORT_SCHEMA_VERSION = "agent-report.v1"
HTML_REPORT_SCHEMA_VERSION = "agent-html-report.v1"
PDF_REPORT_SCHEMA_VERSION = "agent-pdf-report.v1"
REPORT_SHEETS = ["Summary", "Evidence", "Methodology"]
EVIDENCE_HEADERS = [
    "Finding ID",
    "Label",
    "Kind",
    "Value",
    "Unit",
    "Source",
    "Dimension",
    "Warning",
]
_REQUIRED_HTML_SECTIONS = {"summary", "evidence", "methodology"}
_FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
_RUN_ID = re.compile(r"^[0-9a-f]{32}$")


def _required_text(value: Any, name: str, limit: int) -> str:
    if not isinstance(value, str) or not value.strip() or len(value) > limit:
        raise ValueError(f"Agent report {name} is invalid")
    return value


def _optional_text(value: Any, name: str, limit: int) -> str:
    if value is None:
        return ""
    if not isinstance(value, str) or len(value) > limit:
        raise ValueError(f"Agent report {name} is invalid")
    return value


def _canonical_finding(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError("Every Agent report finding must be an object")
    finding_id = _required_text(value.get("finding_id"), "finding_id", 300)
    number = value.get("value")
    if (
        isinstance(number, bool)
        or not isinstance(number, (int, float))
        or not math.isfinite(float(number))
    ):
        raise ValueError(f"Finding {finding_id} has a non-finite numeric value")
    dimension = value.get("dimension", {})
    if not isinstance(dimension, dict) or len(dimension) > 3:
        raise ValueError(f"Finding {finding_id} has an invalid dimension")
    typed_dimension: dict[str, str] = {}
    for key, item in dimension.items():
        if (
            not isinstance(key, str)
            or not key
            or len(key) > 100
            or not isinstance(item, str)
            or len(item) > 200
        ):
            raise ValueError(f"Finding {finding_id} has an invalid dimension")
        typed_dimension[key] = item
    return {
        "finding_id": finding_id,
        "kind": _required_text(value.get("kind"), "finding kind", 100),
        "label": _required_text(value.get("label"), "finding label", 300),
        "value": number,
        "unit": _required_text(value.get("unit"), "finding unit", 100),
        "source": _required_text(value.get("source"), "finding source", 300),
        "dimension": typed_dimension,
        "warning": _optional_text(value.get("warning"), "finding warning", 500),
    }


def verified_agent_report_manifest(run: dict[str, Any]) -> dict[str, Any]:
    """Return the safe, typed projection used by every Agent report format."""

    if not isinstance(run, dict):
        raise ValueError("Agent report history run must be an object")
    run_id = _required_text(run.get("run_id"), "run_id", 32)
    if not _RUN_ID.fullmatch(run_id):
        raise ValueError("Agent report run_id is invalid")
    if run.get("mode") != "agent" or run.get("verifier_status") != "passed":
        raise ValueError("Only verifier-passed Agent history can produce a report")

    raw_findings = run.get("findings")
    if not isinstance(raw_findings, list) or not 1 <= len(raw_findings) <= 48:
        raise ValueError("Agent report requires 1-48 deterministic findings")
    findings = [_canonical_finding(finding) for finding in raw_findings]
    ids = [finding["finding_id"] for finding in findings]
    if len(set(ids)) != len(ids):
        raise ValueError("Agent report finding IDs must be unique")
    if run.get("finding_count") != len(findings):
        raise ValueError("Agent report finding_count does not match the evidence manifest")

    dataset_ref = _required_text(run.get("dataset_ref"), "dataset_ref", 1000)
    normalized_ref = dataset_ref.replace("\\", "/")
    ref_path = PurePosixPath(normalized_ref)
    if ref_path.is_absolute() or ".." in ref_path.parts:
        raise ValueError("Agent report dataset_ref must be workspace-relative")

    raw_tools = run.get("tools")
    if not isinstance(raw_tools, list) or len(raw_tools) > 6:
        raise ValueError("Agent report tool manifest is invalid")
    tools = [_required_text(tool, "tool name", 100) for tool in raw_tools]
    if len(set(tools)) != len(tools):
        raise ValueError("Agent report tool names must be unique")

    return {
        "schema_version": "agent-history-report-source.v1",
        "run_id": run_id,
        "dataset_id": _required_text(run.get("dataset_id"), "dataset_id", 64),
        "dataset_ref": normalized_ref,
        "source_name": _required_text(run.get("source_name"), "source_name", 300),
        "sheet_name": _required_text(run.get("sheet_name"), "sheet_name", 200),
        "question": _required_text(run.get("question"), "question", 4000),
        "run_status": _required_text(run.get("run_status"), "run_status", 40),
        "verifier_status": "passed",
        "created_at": _required_text(run.get("created_at"), "created_at", 100),
        "tools": tools,
        "findings": findings,
        "finding_count": len(findings),
        "business_semantics": "Unverified",
        "evidence_policy": "verified_deterministic_findings_only",
    }


def _manifest_digest(manifest: dict[str, Any]) -> str:
    encoded = json.dumps(
        manifest,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _machine_number(value: int | float) -> str:
    return str(value) if isinstance(value, int) else format(float(value), ".17g")


def _dimension_text(finding: dict[str, Any]) -> str:
    return json.dumps(
        finding.get("dimension", {}), ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )


def _safe_agent_output_path(output_name: str, suffix: str) -> Path:
    if (
        not output_name
        or Path(output_name).name != output_name
        or "/" in output_name
        or "\\" in output_name
    ):
        raise ValueError("Agent report output_name must be a filename, not a path")
    return safe_output_path(output_name, suffix)


def _write_excel(output: Path, manifest: dict[str, Any]) -> None:
    digest = _manifest_digest(manifest)
    with safe_excel_writer(output) as writer:
        workbook = writer.book
        title = workbook.add_format(
            {
                "bold": True,
                "font_size": 20,
                "font_color": "#FFFFFF",
                "bg_color": "#17365D",
            }
        )
        header = workbook.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#2F75B5", "text_wrap": True}
        )
        label = workbook.add_format({"bold": True, "bg_color": "#D9EAF7"})
        wrapped = workbook.add_format({"text_wrap": True, "valign": "top"})
        number = workbook.add_format({"num_format": "0.###############", "align": "right"})
        code = workbook.add_format({"font_name": "Consolas", "font_size": 9})
        warning = workbook.add_format(
            {"font_color": "#7F6000", "bg_color": "#FFF2CC", "text_wrap": True}
        )

        summary = workbook.add_worksheet(REPORT_SHEETS[0])
        summary.hide_gridlines(2)
        summary.merge_range("A1:F1", "LOCAL ANALYTICS COPILOT — VERIFIED AGENT REPORT", title)
        summary_rows = [
            ("Report schema", XLSX_REPORT_SCHEMA_VERSION),
            ("Run ID", manifest["run_id"]),
            ("Evidence manifest SHA-256", digest),
            ("Source file", manifest["source_name"]),
            ("Sheet", manifest["sheet_name"]),
            ("Created at", manifest["created_at"]),
            ("Question", manifest["question"]),
            ("Run status", manifest["run_status"]),
            ("Verifier status", manifest["verifier_status"]),
            ("Deterministic findings", manifest["finding_count"]),
            ("Business semantics", manifest["business_semantics"]),
        ]
        for row, (key, value) in enumerate(summary_rows, start=2):
            summary.write_string(row, 0, str(key), label)
            if isinstance(value, int):
                summary.write_number(row, 1, value, number)
            else:
                summary.write_string(row, 1, str(value), wrapped)
        summary.merge_range(
            "A15:F16",
            "Only verifier-passed deterministic findings are included. Raw rows and model prose "
            "are excluded. Column names do not establish approved business/KPI meaning; "
            "association does not establish causality.",
            warning,
        )
        summary.set_column("A:A", 30)
        summary.set_column("B:F", 30)
        summary.set_row(8, 58)

        evidence = workbook.add_worksheet(REPORT_SHEETS[1])
        evidence.hide_gridlines(2)
        for column, heading in enumerate(EVIDENCE_HEADERS):
            evidence.write_string(0, column, heading, header)
        for row, finding in enumerate(manifest["findings"], start=1):
            evidence.write_string(row, 0, finding["finding_id"], code)
            evidence.write_string(row, 1, finding["label"], wrapped)
            evidence.write_string(row, 2, finding["kind"], wrapped)
            evidence.write_number(row, 3, finding["value"], number)
            evidence.write_string(row, 4, finding["unit"], wrapped)
            evidence.write_string(row, 5, finding["source"], wrapped)
            evidence.write_string(row, 6, _dimension_text(finding), wrapped)
            evidence.write_string(row, 7, finding["warning"], wrapped)
        evidence.autofilter(0, 0, len(manifest["findings"]), len(EVIDENCE_HEADERS) - 1)
        evidence.freeze_panes(1, 1)
        evidence.set_column("A:A", 58)
        evidence.set_column("B:C", 28)
        evidence.set_column("D:E", 18)
        evidence.set_column("F:H", 46)

        methodology = workbook.add_worksheet(REPORT_SHEETS[2])
        methodology.hide_gridlines(2)
        methodology.write_string(0, 0, "Methodology item", header)
        methodology.write_string(0, 1, "Value", header)
        rows = [
            ("Source contract", manifest["schema_version"]),
            ("Evidence policy", manifest["evidence_policy"]),
            ("Tools", ", ".join(manifest["tools"]) or "None"),
            ("Raw data retention", "No raw rows are included in this report"),
            ("Model prose", "Excluded; this report contains deterministic evidence only"),
            ("Business semantics", "Unverified unless separately approved"),
            ("Causality guardrail", "Association is not causality or prediction"),
            ("Manifest SHA-256", digest),
        ]
        for row, (key, value) in enumerate(rows, start=1):
            methodology.write_string(row, 0, key, label)
            methodology.write_string(row, 1, value, wrapped)
        methodology.set_column("A:A", 30)
        methodology.set_column("B:B", 100)


def _same_number(actual: Any, expected: Any) -> bool:
    return (
        isinstance(actual, (int, float))
        and not isinstance(actual, bool)
        and math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-15)
    )


def validate_agent_excel_report(path: Path, run: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, str]] = []
    manifest = verified_agent_report_manifest(run)
    digest = _manifest_digest(manifest)
    try:
        workbook = load_workbook(path, data_only=False, keep_links=False)
    except Exception as exc:
        return {
            "status": "failed",
            "scope": "agent_xlsx_structure_manifest_evidence_and_error_scan",
            "errors": [{"code": "workbook_open_failed", "message": str(exc)[:500]}],
            "finding_count": 0,
            "dashboard_card_count": 0,
            "manifest_sha256": digest,
        }
    if workbook.sheetnames != REPORT_SHEETS:
        errors.append({"code": "invalid_sheet_contract", "message": str(workbook.sheetnames)})

    if "Summary" in workbook.sheetnames:
        summary = workbook["Summary"]
        summary_values = {
            str(summary.cell(row, 1).value): summary.cell(row, 2).value
            for row in range(3, summary.max_row + 1)
            if summary.cell(row, 1).value is not None
        }
        expected_summary = {
            "Report schema": XLSX_REPORT_SCHEMA_VERSION,
            "Run ID": manifest["run_id"],
            "Evidence manifest SHA-256": digest,
            "Verifier status": "passed",
            "Deterministic findings": manifest["finding_count"],
            "Business semantics": "Unverified",
        }
        for key, expected in expected_summary.items():
            if summary_values.get(key) != expected:
                errors.append({"code": "summary_mismatch", "message": key})

    finding_index = {finding["finding_id"]: finding for finding in manifest["findings"]}
    seen: set[str] = set()
    if "Evidence" in workbook.sheetnames:
        evidence = workbook["Evidence"]
        headers = [evidence.cell(1, column).value for column in range(1, 9)]
        if headers != EVIDENCE_HEADERS:
            errors.append({"code": "invalid_evidence_headers", "message": str(headers)})
        for row in range(2, evidence.max_row + 1):
            finding_id = evidence.cell(row, 1).value
            if not isinstance(finding_id, str):
                errors.append({"code": "invalid_evidence_id", "message": str(row)})
                continue
            if finding_id in seen:
                errors.append({"code": "duplicate_evidence_id", "message": finding_id})
                continue
            seen.add(finding_id)
            expected = finding_index.get(finding_id)
            if expected is None:
                errors.append({"code": "unknown_evidence_id", "message": finding_id})
                continue
            actual = [evidence.cell(row, column).value for column in range(2, 9)]
            expected_values = [
                expected["label"],
                expected["kind"],
                expected["value"],
                expected["unit"],
                expected["source"],
                _dimension_text(expected),
                expected["warning"] or None,
            ]
            for column, (found, wanted) in enumerate(zip(actual, expected_values, strict=True), 2):
                matches = (
                    _same_number(found, wanted) if column == 4 else (found or "") == (wanted or "")
                )
                if not matches:
                    errors.append(
                        {"code": "evidence_value_mismatch", "message": f"{finding_id}:{column}"}
                    )
    missing = sorted(set(finding_index) - seen)
    if missing:
        errors.append({"code": "missing_evidence_ids", "message": ", ".join(missing[:20])})

    formula_cells: list[str] = []
    error_cells: list[str] = []
    external_links: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in _FORMULA_ERRORS:
                    error_cells.append(f"{sheet.title}!{cell.coordinate}")
                if cell.data_type == "f":
                    formula_cells.append(f"{sheet.title}!{cell.coordinate}")
                    formula = str(value)
                    if "[" in formula or "://" in formula:
                        external_links.append(f"{sheet.title}!{cell.coordinate}")
    if formula_cells:
        errors.append({"code": "formulas_detected", "message": ", ".join(formula_cells[:20])})
    if error_cells:
        errors.append({"code": "workbook_error_cells", "message": ", ".join(error_cells[:20])})
    if external_links or getattr(workbook, "_external_links", []):
        errors.append(
            {"code": "external_links_detected", "message": ", ".join(external_links[:20])}
        )
    workbook.close()
    return {
        "status": "passed" if not errors else "failed",
        "scope": "agent_xlsx_structure_manifest_evidence_and_error_scan",
        "errors": errors,
        "sheets": REPORT_SHEETS,
        "finding_count": len(seen),
        "dashboard_card_count": 0,
        "formula_count": len(formula_cells),
        "external_links": external_links,
        "manifest_sha256": digest,
    }


def _escape(value: Any) -> str:
    return html.escape(str(value if value is not None else ""), quote=True)


def _html_report(manifest: dict[str, Any]) -> str:
    digest = _manifest_digest(manifest)
    evidence_rows = []
    for finding in manifest["findings"]:
        evidence_rows.append(
            '<tr data-evidence-row="true" '
            f'data-finding-id="{_escape(finding["finding_id"])}" '
            f'data-value="{_escape(_machine_number(finding["value"]))}" '
            f'data-unit="{_escape(finding["unit"])}" '
            f'data-source="{_escape(finding["source"])}">'
            f"<td><code>{_escape(finding['finding_id'])}</code></td>"
            f"<td>{_escape(finding['label'])}</td>"
            f"<td>{_escape(finding['kind'])}</td>"
            f'<td class="number">{_escape(_machine_number(finding["value"]))}</td>'
            f"<td>{_escape(finding['unit'])}</td>"
            f"<td>{_escape(finding['source'])}</td>"
            f"<td><code>{_escape(_dimension_text(finding))}</code></td>"
            f"<td>{_escape(finding['warning'])}</td></tr>"
        )
    tools = ", ".join(manifest["tools"]) or "None"
    return f"""<!doctype html>
<html lang="tr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="lac-report-schema" content="{HTML_REPORT_SCHEMA_VERSION}">
<meta name="lac-manifest-sha256" content="{digest}">
<title>Local Analytics Copilot - Verified Agent Report</title>
<style>
body{{font-family:Arial,sans-serif;color:#263238;margin:32px;line-height:1.45}}
h1,h2{{color:#17365d}} .guardrail{{background:#fff2cc;border:1px solid #bf9000;padding:12px}}
table{{border-collapse:collapse;width:100%;font-size:12px}} th,td{{border:1px solid #cbd7e3;padding:7px;text-align:left;vertical-align:top}}
th{{background:#2f75b5;color:white}} code{{overflow-wrap:anywhere}} .number{{text-align:right}}
</style></head><body>
<h1>Local Analytics Copilot — Verified Agent Report</h1>
<section id="summary"><h2>Summary</h2>
<dl><dt>Run ID</dt><dd>{_escape(manifest["run_id"])}</dd>
<dt>Manifest SHA-256</dt><dd>{digest}</dd>
<dt>Source</dt><dd>{_escape(manifest["source_name"])}</dd>
<dt>Sheet</dt><dd>{_escape(manifest["sheet_name"])}</dd>
<dt>Question</dt><dd>{_escape(manifest["question"])}</dd>
<dt>Verifier</dt><dd>passed</dd><dt>Finding count</dt><dd>{manifest["finding_count"]}</dd></dl>
<p class="guardrail">Only verifier-passed deterministic findings are included. Raw rows and model prose are excluded. Business semantics are Unverified. Association is not causality or prediction.</p>
</section>
<section id="evidence"><h2>Deterministic evidence</h2><table><thead><tr>
<th>Finding ID</th><th>Label</th><th>Kind</th><th>Value</th><th>Unit</th><th>Source</th><th>Dimension</th><th>Warning</th>
</tr></thead><tbody>{"".join(evidence_rows)}</tbody></table></section>
<section id="methodology"><h2>Methodology</h2>
<p>Evidence policy: {_escape(manifest["evidence_policy"])}</p><p>Tools: {_escape(tools)}</p>
<p>Column names do not establish an approved KPI or business definition.</p></section>
</body></html>
"""


class _AgentHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.meta: dict[str, str] = {}
        self.sections: set[str] = set()
        self.evidence_rows: list[dict[str, str]] = []
        self.script_count = 0
        self.external_links: list[str] = []
        self.text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = {key: value or "" for key, value in attrs}
        if tag == "meta" and attributes.get("name"):
            self.meta[attributes["name"]] = attributes.get("content", "")
        if tag == "section" and attributes.get("id"):
            self.sections.add(attributes["id"])
        if tag == "script":
            self.script_count += 1
        for name in ("href", "src"):
            value = attributes.get(name, "").strip()
            if value and not value.startswith(("#", "data:")):
                self.external_links.append(value)
        if tag == "tr" and attributes.get("data-evidence-row") == "true":
            self.evidence_rows.append(attributes)

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.text.append(data.strip())


def validate_agent_html_report(path: Path, run: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, str]] = []
    manifest = verified_agent_report_manifest(run)
    digest = _manifest_digest(manifest)
    try:
        document = path.read_text(encoding="utf-8")
        parser = _AgentHTMLParser()
        parser.feed(document)
        parser.close()
    except Exception as exc:
        return {
            "status": "failed",
            "scope": "agent_html_structure_manifest_evidence_and_external_resource_scan",
            "errors": [{"code": "html_open_failed", "message": str(exc)[:500]}],
            "finding_count": 0,
            "dashboard_card_count": 0,
            "manifest_sha256": digest,
        }
    if not document.lstrip().lower().startswith("<!doctype html>"):
        errors.append({"code": "missing_html_doctype", "message": path.name})
    if parser.meta.get("lac-report-schema") != HTML_REPORT_SCHEMA_VERSION:
        errors.append({"code": "invalid_html_schema", "message": str(parser.meta)})
    if parser.meta.get("lac-manifest-sha256") != digest:
        errors.append({"code": "manifest_digest_mismatch", "message": digest})
    missing_sections = sorted(_REQUIRED_HTML_SECTIONS - parser.sections)
    if missing_sections:
        errors.append({"code": "missing_html_sections", "message": ", ".join(missing_sections)})
    if parser.script_count:
        errors.append({"code": "scripts_detected", "message": str(parser.script_count)})
    if parser.external_links:
        errors.append(
            {"code": "external_links_detected", "message": ", ".join(parser.external_links[:20])}
        )

    finding_index = {finding["finding_id"]: finding for finding in manifest["findings"]}
    seen: set[str] = set()
    for row in parser.evidence_rows:
        finding_id = row.get("data-finding-id", "")
        if finding_id in seen:
            errors.append({"code": "duplicate_evidence_id", "message": finding_id})
            continue
        seen.add(finding_id)
        expected = finding_index.get(finding_id)
        if expected is None:
            errors.append({"code": "unknown_evidence_id", "message": finding_id})
            continue
        try:
            same_value = math.isclose(
                float(row.get("data-value", "nan")),
                float(expected["value"]),
                rel_tol=1e-12,
                abs_tol=1e-15,
            )
        except ValueError:
            same_value = False
        if not same_value:
            errors.append({"code": "evidence_value_mismatch", "message": finding_id})
        if row.get("data-unit") != expected["unit"]:
            errors.append({"code": "evidence_unit_mismatch", "message": finding_id})
        if row.get("data-source") != expected["source"]:
            errors.append({"code": "evidence_source_mismatch", "message": finding_id})
    missing_ids = sorted(set(finding_index) - seen)
    if missing_ids:
        errors.append({"code": "missing_evidence_ids", "message": ", ".join(missing_ids[:20])})
    combined_text = " ".join(parser.text)
    for required in (manifest["run_id"], "Unverified", "Association is not causality"):
        if required not in combined_text:
            errors.append({"code": "required_html_text_missing", "message": required})
    return {
        "status": "passed" if not errors else "failed",
        "scope": "agent_html_structure_manifest_evidence_and_external_resource_scan",
        "errors": errors,
        "finding_count": len(seen),
        "dashboard_card_count": 0,
        "external_links": parser.external_links,
        "manifest_sha256": digest,
    }


def _write_pdf(output: Path, manifest: dict[str, Any]) -> None:
    _register_fonts()
    styles = _pdf_styles()
    digest = _manifest_digest(manifest)
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        str(output),
        pagesize=page_size,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Local Analytics Copilot - Verified Agent Report",
        author="Local Analytics Copilot",
        subject=PDF_REPORT_SCHEMA_VERSION,
    )

    def page_frame(canvas, doc) -> None:
        canvas.saveState()
        canvas.setTitle("Local Analytics Copilot - Verified Agent Report")
        canvas.setAuthor("Local Analytics Copilot")
        canvas.setSubject(PDF_REPORT_SCHEMA_VERSION)
        canvas.setCreator("Local Analytics Copilot deterministic report engine")
        canvas.setKeywords(f"{PDF_REPORT_SCHEMA_VERSION};lac-manifest-sha256={digest}")
        canvas.setStrokeColor(colors.HexColor("#D9E2F3"))
        canvas.line(12 * mm, 10 * mm, page_size[0] - 12 * mm, 10 * mm)
        canvas.setFont("LACVera", 7)
        canvas.setFillColor(colors.HexColor("#5D6B75"))
        canvas.drawString(12 * mm, 6.2 * mm, PDF_REPORT_SCHEMA_VERSION)
        canvas.drawRightString(page_size[0] - 12 * mm, 6.2 * mm, f"Page {doc.page}")
        canvas.restoreState()

    summary_rows = [
        ["Run ID", manifest["run_id"], "Verifier", "passed"],
        ["Source", manifest["source_name"], "Sheet", manifest["sheet_name"]],
        ["Findings", str(manifest["finding_count"]), "Business semantics", "Unverified"],
    ]
    summary_table = Table(
        [[_pdf_paragraph(cell, styles["body"]) for cell in row] for row in summary_rows],
        colWidths=[32 * mm, 72 * mm, 38 * mm, 71 * mm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "LACVera"),
                ("FONTNAME", (0, 0), (0, -1), "LACVeraBold"),
                ("FONTNAME", (2, 0), (2, -1), "LACVeraBold"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF5FB")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD7E3")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story: list[Any] = [
        _pdf_paragraph("Local Analytics Copilot", styles["title"]),
        _pdf_paragraph(
            "Verified Agent Report - deterministic evidence only; no raw rows or model prose.",
            styles["subtitle"],
        ),
        _pdf_paragraph("Summary", styles["heading"]),
        summary_table,
        Spacer(1, 3 * mm),
        _pdf_paragraph("Question", styles["heading"]),
        _pdf_paragraph(manifest["question"], styles["body"]),
        _pdf_paragraph(
            "Business semantics are Unverified. Column names do not establish an approved KPI. "
            "Association is not causality or prediction.",
            styles["warning"],
        ),
        PageBreak(),
        _pdf_paragraph("Deterministic evidence", styles["heading"]),
    ]
    evidence_data = [["Finding ID", "Label", "Kind", "Value", "Unit", "Source", "Dimension"]]
    for finding in manifest["findings"]:
        evidence_data.append(
            [
                finding["finding_id"],
                finding["label"],
                finding["kind"],
                _machine_number(finding["value"]),
                finding["unit"],
                finding["source"],
                _dimension_text(finding),
            ]
        )
    evidence_table = LongTable(
        [
            [
                _pdf_paragraph(cell, styles["small_bold"] if index == 0 else styles["small"])
                for cell in row
            ]
            for index, row in enumerate(evidence_data)
        ],
        colWidths=[50 * mm, 37 * mm, 27 * mm, 22 * mm, 23 * mm, 45 * mm, 39 * mm],
        repeatRows=1,
        splitByRow=1,
    )
    evidence_table.setStyle(_pdf_table_style(font_size=6.5))
    story.extend(
        [
            evidence_table,
            PageBreak(),
            _pdf_paragraph("Methodology and verification", styles["heading"]),
        ]
    )
    methodology = [
        ("PDF report schema", PDF_REPORT_SCHEMA_VERSION),
        ("Source contract", manifest["schema_version"]),
        ("Evidence manifest SHA-256", digest),
        ("Evidence policy", manifest["evidence_policy"]),
        ("Deterministic tools", ", ".join(manifest["tools"]) or "None"),
        ("Raw data retention", "No raw rows are included"),
        ("Model prose", "Excluded from this evidence report"),
    ]
    methodology_table = Table(
        [
            [_pdf_paragraph(key, styles["body"]), _pdf_paragraph(value, styles["body"])]
            for key, value in methodology
        ],
        colWidths=[50 * mm, 163 * mm],
    )
    methodology_table.setStyle(_pdf_table_style(header=False, font_size=8.2))
    story.extend([methodology_table, Spacer(1, 3 * mm), _pdf_paragraph(digest, styles["small"])])
    document.build(story, onFirstPage=page_frame, onLaterPages=page_frame)


def _compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value if value is not None else ""))


def validate_agent_pdf_report(path: Path, run: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, str]] = []
    manifest = verified_agent_report_manifest(run)
    digest = _manifest_digest(manifest)
    try:
        reader = PdfReader(path)
        if reader.is_encrypted:
            raise ValueError("Encrypted reports are not accepted")
        pages = list(reader.pages)
        text = "\n".join(page.extract_text() or "" for page in pages)
        metadata = reader.metadata or {}
    except Exception as exc:
        return {
            "status": "failed",
            "scope": "agent_pdf_structure_manifest_evidence_and_external_link_scan",
            "errors": [{"code": "pdf_open_failed", "message": str(exc)[:500]}],
            "page_count": 0,
            "finding_count": 0,
            "dashboard_card_count": 0,
            "manifest_sha256": digest,
        }
    if not pages:
        errors.append({"code": "empty_pdf", "message": path.name})
    if str(metadata.get("/Title", "")) != "Local Analytics Copilot - Verified Agent Report":
        errors.append({"code": "invalid_pdf_title", "message": str(metadata.get("/Title", ""))})
    if str(metadata.get("/Subject", "")) != PDF_REPORT_SCHEMA_VERSION:
        errors.append({"code": "invalid_pdf_schema", "message": str(metadata.get("/Subject", ""))})
    if f"lac-manifest-sha256={digest}" not in str(metadata.get("/Keywords", "")):
        errors.append({"code": "manifest_digest_mismatch", "message": digest})
    external_links: list[str] = []
    for page_number, page in enumerate(pages, start=1):
        if float(page.mediabox.width) <= 0 or float(page.mediabox.height) <= 0:
            errors.append({"code": "invalid_page_box", "message": str(page_number)})
        if page.get("/Annots"):
            external_links.append(f"page:{page_number}")
    if external_links:
        errors.append({"code": "external_links_detected", "message": ", ".join(external_links)})

    compact = _compact_text(text)
    for required in (manifest["run_id"], "Unverified", digest):
        if _compact_text(required) not in compact:
            errors.append({"code": "required_pdf_text_missing", "message": required})
    found_ids = 0
    for finding in manifest["findings"]:
        if _compact_text(finding["finding_id"]) not in compact:
            errors.append({"code": "missing_evidence_id", "message": finding["finding_id"]})
            continue
        found_ids += 1
        for field in ("unit", "source"):
            if _compact_text(finding[field]) not in compact:
                errors.append(
                    {"code": f"evidence_{field}_missing", "message": finding["finding_id"]}
                )
        if _compact_text(_machine_number(finding["value"])) not in compact:
            errors.append({"code": "evidence_value_missing", "message": finding["finding_id"]})
    return {
        "status": "passed" if not errors else "failed",
        "scope": "agent_pdf_structure_manifest_evidence_and_external_link_scan",
        "errors": errors,
        "page_count": len(pages),
        "finding_count": found_ids,
        "dashboard_card_count": 0,
        "external_links": external_links,
        "manifest_sha256": digest,
    }


def _report_result(
    output: Path,
    manifest: dict[str, Any],
    schema_version: str,
    media_type: str,
    verification: dict[str, Any],
) -> dict[str, Any]:
    if verification["status"] != "passed":
        output.unlink(missing_ok=True)
        raise RuntimeError(f"Agent report verification failed: {verification['errors'][:3]}")
    settings = get_settings()
    relative_output = str(output.resolve().relative_to(settings.workspace.resolve()))
    audit(
        settings.logs_dir,
        "agent_verified_report",
        output=relative_output,
        run_id=manifest["run_id"],
        schema=schema_version,
        findings=verification["finding_count"],
    )
    return {
        "status": "created",
        "schema_version": schema_version,
        "output": relative_output,
        "filename": output.name,
        "media_type": media_type,
        "verification": verification,
    }


def create_agent_excel_report(
    run: dict[str, Any], output_name: str = "agent_report.xlsx"
) -> dict[str, Any]:
    manifest = verified_agent_report_manifest(run)
    output = _safe_agent_output_path(output_name, ".xlsx")
    _write_excel(output, manifest)
    verification = validate_agent_excel_report(output, run)
    return _report_result(
        output,
        manifest,
        XLSX_REPORT_SCHEMA_VERSION,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        verification,
    )


def create_agent_html_report(
    run: dict[str, Any], output_name: str = "agent_report.html"
) -> dict[str, Any]:
    manifest = verified_agent_report_manifest(run)
    output = _safe_agent_output_path(output_name, ".html")
    output.write_text(_html_report(manifest), encoding="utf-8", newline="\n")
    verification = validate_agent_html_report(output, run)
    return _report_result(
        output,
        manifest,
        HTML_REPORT_SCHEMA_VERSION,
        "text/html; charset=utf-8",
        verification,
    )


def create_agent_pdf_report(
    run: dict[str, Any], output_name: str = "agent_report.pdf"
) -> dict[str, Any]:
    manifest = verified_agent_report_manifest(run)
    output = _safe_agent_output_path(output_name, ".pdf")
    _write_pdf(output, manifest)
    verification = validate_agent_pdf_report(output, run)
    return _report_result(
        output,
        manifest,
        PDF_REPORT_SCHEMA_VERSION,
        "application/pdf",
        verification,
    )


def create_agent_report(
    run: dict[str, Any], report_format: AgentReportFormat, output_name: str | None = None
) -> dict[str, Any]:
    factories = {
        "xlsx": (create_agent_excel_report, "agent_report.xlsx"),
        "html": (create_agent_html_report, "agent_report.html"),
        "pdf": (create_agent_pdf_report, "agent_report.pdf"),
    }
    factory, default_name = factories[report_format]
    return factory(run, output_name or default_name)
