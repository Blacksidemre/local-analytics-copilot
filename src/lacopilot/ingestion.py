from __future__ import annotations

import csv
import io
import re
import statistics
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path, PurePosixPath
from typing import Any

import pandas as pd

from lacopilot.config import get_settings
from lacopilot.security import validate_file_size

SUPPORTED_TABLE_EXTENSIONS = {
    ".csv",
    ".xlsx",
    ".xlsm",
    ".xls",
    ".parquet",
    ".pq",
    ".json",
    ".jsonl",
}
_DELIMITERS = (",", ";", "\t", "|")
_DELIMITER_LABELS = {",": "comma", ";": "semicolon", "\t": "tab", "|": "pipe"}
_UTF_BOMS = (
    (b"\xef\xbb\xbf", "utf-8-sig"),
    (b"\xff\xfe\x00\x00", "utf-32"),
    (b"\x00\x00\xfe\xff", "utf-32"),
    (b"\xff\xfe", "utf-16"),
    (b"\xfe\xff", "utf-16"),
)
_INTEGER = re.compile(r"^[+-]?\d+$")
_DECIMAL_COMMA = re.compile(r"^[+-]?(?:\d{1,3}(?:\.\d{3})+|\d+),\d+$")
_DECIMAL_DOT = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d+$")


class IngestionError(ValueError):
    """Stable, user-facing error raised by deterministic file ingestion."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        hint: str | None = None,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.hint = hint
        self.details = details or {}

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "message": self.message,
            "hint": self.hint,
            "details": self.details,
        }


@dataclass(frozen=True)
class HeaderRename:
    index: int
    original: str
    normalized: str


@dataclass(frozen=True)
class CsvDialect:
    encoding: str
    delimiter: str
    delimiter_name: str
    quotechar: str
    decimal_separator: str
    thousands_separator: str | None
    expected_columns: int
    headers: tuple[str, ...]
    sampled_rows: int
    consistency: float
    confidence: str
    header_renames: tuple[HeaderRename, ...] = ()
    warnings: tuple[str, ...] = ()

    def as_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["header_renames"] = [asdict(item) for item in self.header_renames]
        return payload


@dataclass
class TableReadResult:
    dataframe: pd.DataFrame
    metadata: dict[str, Any] = field(default_factory=dict)


def _sample_bytes(path: Path, limit: int = 256 * 1024) -> bytes:
    with path.open("rb") as source:
        return source.read(limit)


def _decode_csv_sample(raw: bytes) -> tuple[str, str, list[str]]:
    if not raw:
        raise IngestionError("empty_file", "Dosya boş.", hint="Veri içeren bir dosya seçin.")
    if b"\x00" in raw and not any(raw.startswith(bom) for bom, _ in _UTF_BOMS):
        raise IngestionError(
            "binary_csv",
            "CSV olarak seçilen dosya metin dosyası görünmüyor.",
            hint="Dosya uzantısını ve gerçek dosya türünü kontrol edin.",
        )

    warnings: list[str] = []
    for bom, encoding in _UTF_BOMS:
        if raw.startswith(bom):
            try:
                return raw.decode(encoding), encoding, warnings
            except UnicodeDecodeError as exc:
                raise IngestionError(
                    "encoding_error",
                    f"Dosya {encoding} olarak çözülemedi.",
                    details={"offset": exc.start},
                ) from exc

    for encoding in ("utf-8", "cp1254", "cp1252"):
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        if encoding != "utf-8":
            warnings.append(f"UTF-8 yerine {encoding} encoding algılandı.")
        return text, encoding, warnings

    warnings.append("Encoding kesin belirlenemedi; latin-1 geri dönüşü kullanıldı.")
    return raw.decode("latin-1"), "latin-1", warnings


def _sniff_quotechar(text: str) -> tuple[str | None, str]:
    try:
        dialect = csv.Sniffer().sniff(text, delimiters="".join(_DELIMITERS))
        quotechar = dialect.quotechar if dialect.quotechar in {'"', "'"} else '"'
        return dialect.delimiter, quotechar
    except csv.Error:
        return None, '"'


def _candidate_rows(text: str, delimiter: str, quotechar: str) -> tuple[list[list[str]], bool]:
    rows: list[list[str]] = []
    try:
        reader = csv.reader(
            io.StringIO(text),
            delimiter=delimiter,
            quotechar=quotechar,
            doublequote=True,
            strict=True,
        )
        for row in reader:
            if not row or not any(cell.strip() for cell in row):
                continue
            rows.append(row)
            if len(rows) >= 80:
                break
    except csv.Error:
        return rows, False
    return rows, True


def _score_delimiter(
    rows: list[list[str]], parsed_cleanly: bool, delimiter: str, sniffed: str | None
) -> tuple[float, int, float]:
    if not rows:
        return -100.0, 1, 0.0
    widths = [len(row) for row in rows]
    mode_width = statistics.multimode(widths)[0]
    consistency = widths.count(mode_width) / len(widths)
    multi_column = mode_width > 1
    score = consistency * 8.0
    score += min(mode_width, 20) * 0.12 if multi_column else -2.0
    score += 1.5 if sniffed == delimiter else 0.0
    score += 0.5 if parsed_cleanly else -4.0
    if len(rows) >= 3 and len(set(widths[: min(20, len(widths))])) == 1:
        score += 1.0
    return score, mode_width, consistency


def _detect_number_separators(rows: list[list[str]]) -> tuple[str, str | None]:
    comma_decimal = 0
    dot_decimal = 0
    comma_thousands = 0
    dot_thousands = 0
    for row in rows[1:60]:
        for raw in row:
            value = raw.strip().replace("\u00a0", "").replace(" ", "")
            if not value or _INTEGER.fullmatch(value):
                continue
            if _DECIMAL_COMMA.fullmatch(value):
                comma_decimal += 1
                if "." in value:
                    dot_thousands += 1
            elif _DECIMAL_DOT.fullmatch(value):
                dot_decimal += 1
                if "," in value:
                    comma_thousands += 1
    if comma_decimal > dot_decimal and comma_decimal >= 2:
        return ",", "." if dot_thousands >= 1 else None
    return ".", "," if comma_thousands >= 1 else None


def _normalize_headers(values: list[Any]) -> tuple[list[str], tuple[HeaderRename, ...]]:
    clean: list[str] = []
    renames: list[HeaderRename] = []
    next_suffix: dict[str, int] = {}
    used: set[str] = set()
    for index, value in enumerate(values):
        original = "" if value is None else str(value).strip()
        base = original
        if not base or base.lower().startswith("unnamed:"):
            base = f"column_{index + 1}"
        normalized = base
        suffix = next_suffix.get(base, 2)
        while normalized in used:
            normalized = f"{base}_{suffix}"
            suffix += 1
        next_suffix[base] = suffix
        used.add(normalized)
        clean.append(normalized)
        if normalized != original:
            renames.append(HeaderRename(index=index, original=original, normalized=normalized))
    return clean, tuple(renames)


def detect_csv_dialect(path: Path) -> CsvDialect:
    raw = _sample_bytes(path)
    text, encoding, warnings = _decode_csv_sample(raw)
    sniffed, quotechar = _sniff_quotechar(text)
    candidates: list[tuple[float, str, int, float, list[list[str]]]] = []
    for delimiter in _DELIMITERS:
        rows, clean = _candidate_rows(text, delimiter, quotechar)
        score, width, consistency = _score_delimiter(rows, clean, delimiter, sniffed)
        candidates.append((score, delimiter, width, consistency, rows))
    candidates.sort(key=lambda item: item[0], reverse=True)
    _, delimiter, width, consistency, rows = candidates[0]

    if width == 1:
        delimiter = sniffed if sniffed in _DELIMITERS else ","
        rows, _ = _candidate_rows(text, delimiter, quotechar)
        width = 1
        consistency = 1.0 if rows else 0.0
        warnings.append("Tek sütunlu dosya algılandı; delimiter doğrulanamadığı için güven düşük.")

    if not rows:
        raise IngestionError(
            "empty_csv",
            "CSV içinde okunabilir satır bulunamadı.",
            hint="Dosyanın encoding ve delimiter ayarlarını kontrol edin.",
        )

    decimal, thousands = _detect_number_separators(rows)
    headers, renames = _normalize_headers(rows[0])
    width = max(width, len(headers))
    if consistency >= 0.98 and width > 1:
        confidence = "high"
    elif consistency >= 0.85 and width > 1:
        confidence = "medium"
        warnings.append("Bazı örnek satırların sütun sayısı farklı; tam okuma doğrulanacak.")
    else:
        confidence = "low"

    return CsvDialect(
        encoding=encoding,
        delimiter=delimiter,
        delimiter_name=_DELIMITER_LABELS[delimiter],
        quotechar=quotechar,
        decimal_separator=decimal,
        thousands_separator=thousands,
        expected_columns=width,
        headers=tuple(headers),
        sampled_rows=len(rows),
        consistency=round(consistency, 4),
        confidence=confidence,
        header_renames=renames,
        warnings=tuple(warnings),
    )


def _validate_csv_structure(path: Path, dialect: CsvDialect) -> None:
    try:
        with path.open("r", encoding=dialect.encoding, newline="") as source:
            reader = csv.reader(
                source,
                delimiter=dialect.delimiter,
                quotechar=dialect.quotechar,
                doublequote=True,
                strict=True,
            )
            logical_row = 0
            for row in reader:
                if not row or not any(cell.strip() for cell in row):
                    continue
                logical_row += 1
                if len(row) != dialect.expected_columns:
                    raise IngestionError(
                        "inconsistent_csv_row",
                        "CSV satırlarında farklı sütun sayıları bulundu; dosya eksik okunmadı.",
                        hint="Belirtilen satırdaki delimiter ve quote karakterlerini kontrol edin.",
                        details={
                            "logical_row": logical_row,
                            "physical_line": reader.line_num,
                            "expected_columns": dialect.expected_columns,
                            "actual_columns": len(row),
                            "detection": dialect.as_dict(),
                        },
                    )
    except IngestionError:
        raise
    except (csv.Error, UnicodeDecodeError) as exc:
        raise IngestionError(
            "malformed_csv",
            "CSV quote veya encoding yapısı geçerli değil; dosya eksik okunmadı.",
            details={"parser_error": str(exc)[:500], "detection": dialect.as_dict()},
        ) from exc


def _read_csv(path: Path, nrows: int | None) -> TableReadResult:
    dialect = detect_csv_dialect(path)
    _validate_csv_structure(path, dialect)
    try:
        dataframe = pd.read_csv(
            path,
            nrows=nrows,
            encoding=dialect.encoding,
            sep=dialect.delimiter,
            quotechar=dialect.quotechar,
            decimal=dialect.decimal_separator,
            thousands=dialect.thousands_separator,
            engine="python",
            on_bad_lines="error",
        )
    except (csv.Error, pd.errors.ParserError, UnicodeDecodeError) as exc:
        raise IngestionError(
            "malformed_csv",
            "CSV satır yapısı tutarlı değil; dosya sessizce eksik okunmadı.",
            hint="Hatalı satırı, quote karakterlerini veya delimiter seçimini kontrol edin.",
            details={"parser_error": str(exc)[:500], "detection": dialect.as_dict()},
        ) from exc
    if dataframe.shape[1] != dialect.expected_columns:
        raise IngestionError(
            "csv_shape_mismatch",
            "CSV ön incelemesi ile tam okuma farklı sütun sayıları üretti.",
            hint="Dosyada bozuk quote, karışık delimiter veya düzensiz satır olabilir.",
            details={
                "expected_columns": dialect.expected_columns,
                "parsed_columns": int(dataframe.shape[1]),
                "detection": dialect.as_dict(),
            },
        )
    headers = list(dialect.headers)
    renames = dialect.header_renames
    dataframe.columns = headers
    metadata = dialect.as_dict()
    if renames:
        metadata["header_renames"] = [asdict(item) for item in renames]
    return TableReadResult(
        dataframe=dataframe,
        metadata={"format": "csv", "csv": metadata},
    )


def _row_values(worksheet: Any, row_index: int, max_columns: int) -> list[Any]:
    row = next(
        worksheet.iter_rows(
            min_row=row_index,
            max_row=row_index,
            min_col=1,
            max_col=max_columns,
            values_only=True,
        ),
        (),
    )
    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return values


def _detect_excel_header(worksheet: Any) -> int:
    max_rows = min(int(worksheet.max_row or 1), 25)
    max_columns = min(max(int(worksheet.max_column or 1), 1), 2000)
    candidates: list[tuple[float, int]] = []
    for row_index in range(1, max_rows + 1):
        values = _row_values(worksheet, row_index, max_columns)
        nonempty = [value for value in values if value is not None and str(value).strip()]
        if len(nonempty) < 2:
            continue
        labels = [str(value).strip() for value in nonempty]
        uniqueness = len(set(labels)) / len(labels)
        text_ratio = sum(not isinstance(value, (int, float)) for value in nonempty) / len(nonempty)
        next_fill: list[float] = []
        for next_index in range(row_index + 1, min(row_index + 6, max_rows + 1)):
            following = _row_values(worksheet, next_index, max_columns)
            filled = sum(bool(value is not None and str(value).strip()) for value in following)
            # Compare all candidates against the worksheet width.  Using the
            # trimmed candidate-row width can give data rows a score above 1.0
            # and incorrectly select them as the header.
            next_fill.append(filled / max(max_columns, 1))
        continuation = statistics.median(next_fill) if next_fill else 0.0
        score = uniqueness * 3.0 + text_ratio * 1.4 + continuation * 4.0
        score += min(len(nonempty), 30) * 0.04
        score -= (row_index - 1) * 0.015
        candidates.append((score, row_index))
    if not candidates:
        return 1
    return max(candidates)[1]


def validate_excel_archive(path: Path, limit_bytes: int | None = None) -> dict[str, int]:
    settings = get_settings()
    limit = limit_bytes or int(settings.max_excel_uncompressed_mb) * 1024 * 1024
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > settings.max_excel_archive_entries:
                raise IngestionError(
                    "excel_archive_too_many_entries",
                    "Excel paketi güvenli giriş sayısı sınırını aşıyor.",
                    details={
                        "entries": len(entries),
                        "max_entries": settings.max_excel_archive_entries,
                    },
                )
            total = 0
            for entry in entries:
                parts = PurePosixPath(entry.filename).parts
                if entry.filename.startswith("/") or ".." in parts:
                    raise IngestionError(
                        "unsafe_excel_archive_path",
                        "Excel paketi güvenli olmayan bir iç dosya yolu içeriyor.",
                    )
                if entry.flag_bits & 0x1:
                    raise IngestionError(
                        "encrypted_workbook",
                        "Parola korumalı Excel dosyaları analiz için desteklenmiyor.",
                        hint="Yetkiliyseniz parolayı Excel'de kaldırıp yeni bir kopya kaydedin.",
                    )
                total += int(entry.file_size)
                if total > limit:
                    raise IngestionError(
                        "excel_decompression_limit",
                        "Excel dosyası açıldığında güvenli bellek sınırını aşıyor.",
                        hint="Gerekli sayfayı CSV veya Parquet olarak dışa aktarın.",
                        details={"declared_uncompressed_bytes": total, "max_bytes": limit},
                    )
    except IngestionError:
        raise
    except (OSError, zipfile.BadZipFile) as exc:
        raise IngestionError(
            "invalid_workbook_archive",
            "Excel dosyasının ZIP paketi geçerli değil.",
            details={"reader_error": str(exc)[:500]},
        ) from exc
    return {"entries": len(entries), "declared_uncompressed_bytes": total}


def inspect_workbook(path: Path) -> dict[str, Any]:
    archive = validate_excel_archive(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngestionError(
            "excel_dependency_missing",
            "Excel desteği için openpyxl kurulu değil.",
            hint="Kurulumu yeniden çalıştırın.",
        ) from exc
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise IngestionError(
            "invalid_workbook",
            "Excel çalışma kitabı açılamadı.",
            hint="Dosyanın bozuk, parola korumalı veya yanlış uzantılı olmadığını kontrol edin.",
            details={"reader_error": str(exc)[:500]},
        ) from exc
    try:
        sheets = []
        for index, worksheet in enumerate(workbook.worksheets):
            header_row = _detect_excel_header(worksheet)
            headers = _row_values(
                worksheet,
                header_row,
                min(max(int(worksheet.max_column or 1), 1), 2000),
            )
            normalized, renames = _normalize_headers(headers)
            sheets.append(
                {
                    "index": index,
                    "name": worksheet.title,
                    "header_row": header_row,
                    "estimated_rows": max(int(worksheet.max_row or 0) - header_row, 0),
                    "estimated_columns": len(normalized),
                    "columns": normalized,
                    "header_renames": [asdict(item) for item in renames],
                    "empty": not normalized or int(worksheet.max_row or 0) <= header_row,
                }
            )
    finally:
        workbook.close()
    if not sheets:
        raise IngestionError("no_sheets", "Excel dosyasında çalışma sayfası bulunamadı.")
    return {
        "format": path.suffix.lower().lstrip("."),
        "sheets": sheets,
        "archive": archive,
    }


def _resolve_sheet(manifest: dict[str, Any], sheet_name: str | int | None) -> dict[str, Any]:
    sheets = manifest["sheets"]
    requested: str | int = 0 if sheet_name is None else sheet_name
    if isinstance(requested, str) and requested.isdigit():
        requested = int(requested)
    if isinstance(requested, int):
        if 0 <= requested < len(sheets):
            return sheets[requested]
    else:
        for sheet in sheets:
            if sheet["name"] == requested:
                return sheet
    raise IngestionError(
        "sheet_not_found",
        f"Excel sayfası bulunamadı: {sheet_name}",
        details={"available_sheets": [sheet["name"] for sheet in sheets]},
    )


def _read_excel(path: Path, sheet_name: str | int | None, nrows: int | None) -> TableReadResult:
    if path.suffix.lower() == ".xls":
        try:
            dataframe = pd.read_excel(path, sheet_name=sheet_name or 0, nrows=nrows)
        except Exception as exc:
            raise IngestionError(
                "invalid_workbook",
                "Eski .xls dosyası okunamadı.",
                hint="Dosyayı .xlsx olarak kaydedip yeniden deneyin.",
                details={"reader_error": str(exc)[:500]},
            ) from exc
        headers, renames = _normalize_headers(list(dataframe.columns))
        dataframe.columns = headers
        return TableReadResult(
            dataframe=dataframe,
            metadata={
                "format": "xls",
                "excel": {
                    "selected_sheet": str(sheet_name or 0),
                    "header_row": 1,
                    "header_renames": [asdict(item) for item in renames],
                    "legacy_format": True,
                },
            },
        )

    manifest = inspect_workbook(path)
    sheet = _resolve_sheet(manifest, sheet_name)
    try:
        dataframe = pd.read_excel(
            path,
            sheet_name=sheet["name"],
            header=sheet["header_row"] - 1,
            nrows=nrows,
            engine="openpyxl",
        )
    except Exception as exc:
        raise IngestionError(
            "excel_parse_error",
            f"Excel sayfası okunamadı: {sheet['name']}",
            details={"reader_error": str(exc)[:500]},
        ) from exc
    dataframe = dataframe.dropna(how="all")
    headers, renames = _normalize_headers(list(dataframe.columns))
    dataframe.columns = headers
    if not len(headers):
        raise IngestionError("empty_sheet", f"Excel sayfasında sütun bulunamadı: {sheet['name']}")
    return TableReadResult(
        dataframe=dataframe,
        metadata={
            "format": manifest["format"],
            "excel": {
                "selected_sheet": sheet["name"],
                "selected_sheet_index": sheet["index"],
                "header_row": sheet["header_row"],
                "available_sheets": [item["name"] for item in manifest["sheets"]],
                "header_renames": [asdict(item) for item in renames],
            },
        },
    )


def read_table(
    path: Path,
    sheet_name: str | int | None = 0,
    nrows: int | None = None,
) -> TableReadResult:
    settings = get_settings()
    validate_file_size(path, settings.max_file_mb)
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_TABLE_EXTENSIONS:
        raise IngestionError(
            "unsupported_type",
            f"Desteklenmeyen dosya tipi: {suffix or '(uzantı yok)'}",
            details={"supported": sorted(SUPPORTED_TABLE_EXTENSIONS)},
        )
    if suffix == ".csv":
        result = _read_csv(path, nrows)
    elif suffix in {".xlsx", ".xlsm", ".xls"}:
        result = _read_excel(path, sheet_name, nrows)
    elif suffix in {".parquet", ".pq"}:
        try:
            dataframe = pd.read_parquet(path)
        except Exception as exc:
            raise IngestionError(
                "parquet_parse_error",
                "Parquet dosyası okunamadı.",
                hint="PyArrow kurulumunu ve dosyanın bütünlüğünü kontrol edin.",
                details={"reader_error": str(exc)[:500]},
            ) from exc
        if nrows is not None:
            dataframe = dataframe.head(nrows)
        headers, renames = _normalize_headers(list(dataframe.columns))
        dataframe.columns = headers
        result = TableReadResult(
            dataframe=dataframe,
            metadata={
                "format": "parquet",
                "parquet": {"header_renames": [asdict(item) for item in renames]},
            },
        )
    else:
        try:
            dataframe = pd.read_json(path, lines=suffix == ".jsonl")
        except Exception as exc:
            raise IngestionError(
                "json_parse_error",
                "JSON tablosu okunamadı.",
                hint="Dosyanın JSON object listesi veya JSON Lines biçiminde olduğunu kontrol edin.",
                details={"reader_error": str(exc)[:500]},
            ) from exc
        if nrows is not None:
            dataframe = dataframe.head(nrows)
        headers, renames = _normalize_headers(list(dataframe.columns))
        dataframe.columns = headers
        result = TableReadResult(
            dataframe=dataframe,
            metadata={
                "format": suffix.lstrip("."),
                "json": {"header_renames": [asdict(item) for item in renames]},
            },
        )
    result.metadata.update(
        {
            "source_name": path.name,
            "size_bytes": path.stat().st_size,
            "parser": "lac-deterministic-data-bridge",
            "parser_version": 1,
        }
    )
    return result


def source_manifest(path: Path) -> dict[str, Any]:
    settings = get_settings()
    validate_file_size(path, settings.max_file_mb)
    suffix = path.suffix.lower()
    base = {
        "source_name": path.name,
        "size_bytes": path.stat().st_size,
        "extension": suffix,
        "parser": "lac-deterministic-data-bridge",
        "parser_version": 1,
    }
    if suffix == ".csv":
        return base | {"format": "csv", "csv": detect_csv_dialect(path).as_dict()}
    if suffix in {".xlsx", ".xlsm"}:
        return base | inspect_workbook(path)
    if suffix == ".xls":
        return base | {"format": "xls", "legacy_format": True}
    if suffix in {".parquet", ".pq"}:
        return base | {"format": "parquet"}
    if suffix in {".json", ".jsonl"}:
        return base | {"format": suffix.lstrip(".")}
    raise IngestionError(
        "unsupported_type",
        f"Desteklenmeyen dosya tipi: {suffix or '(uzantı yok)'}",
        details={"supported": sorted(SUPPORTED_TABLE_EXTENSIONS)},
    )
