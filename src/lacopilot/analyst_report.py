from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from lacopilot.audit import audit
from lacopilot.config import get_settings
from lacopilot.tools.common import safe_excel_writer, safe_output_path

REPORT_SCHEMA_VERSION = "analyst-report.v1"
REPORT_SHEETS = ["Executive Dashboard", "Associations", "Evidence", "Methodology"]
EVIDENCE_HEADERS = [
    "Finding ID",
    "Label",
    "Kind",
    "Value",
    "Unit",
    "Source",
    "Dimension",
    "Warning",
]
_DIRECT_EVIDENCE_FORMULA = re.compile(r"^='Evidence'!\$([A-H])\$(\d+)$")
_FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def analyst_manifest_sha256(payload: dict[str, Any]) -> str:
    """Bind every Analyst report format to the same verified evidence manifest."""
    manifest = {
        "schema_version": payload["schema_version"],
        "target_semantics": payload["target_semantics"],
        "kpi_selection": payload["kpi_selection"],
        "multiple_testing": payload["multiple_testing"],
        "analyses": payload["analyses"],
        "findings": payload["findings"],
        "dashboard": payload["dashboard"],
        "interpretation": payload["interpretation"],
        "verification": payload["verification"],
    }
    encoded = json.dumps(
        manifest,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _finding_index(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        finding["finding_id"]: finding
        for finding in payload.get("findings", [])
        if isinstance(finding, dict) and isinstance(finding.get("finding_id"), str)
    }


def _text(value: Any, limit: int = 32_000) -> str:
    return str(value if value is not None else "")[:limit]


def _write_text(worksheet, row: int, column: int, value: Any, cell_format=None) -> None:
    worksheet.write_string(row, column, _text(value), cell_format)


def _write_report_workbook(output: Path, payload: dict[str, Any]) -> None:
    findings = payload["findings"]
    finding_index = _finding_index(payload)
    evidence_rows = {finding["finding_id"]: row for row, finding in enumerate(findings, start=2)}
    analyses_by_effect = {
        analysis["finding_ids"]["effect"]: analysis for analysis in payload["analyses"]
    }

    with safe_excel_writer(output) as writer:
        workbook = writer.book
        workbook.set_properties(
            {
                "keywords": (
                    f"{REPORT_SCHEMA_VERSION};"
                    f"lac-manifest-sha256={analyst_manifest_sha256(payload)}"
                )
            }
        )
        title = workbook.add_format(
            {
                "bold": True,
                "font_size": 20,
                "font_color": "#FFFFFF",
                "bg_color": "#17365D",
                "align": "left",
                "valign": "vcenter",
            }
        )
        subtitle = workbook.add_format(
            {"font_color": "#44546A", "bg_color": "#D9EAF7", "text_wrap": True}
        )
        section = workbook.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#1F4E78"}
        )
        header = workbook.add_format(
            {
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": "#2F75B5",
                "bottom": 1,
                "bottom_color": "#D9E2F3",
                "text_wrap": True,
            }
        )
        note = workbook.add_format(
            {"font_color": "#7F6000", "bg_color": "#FFF2CC", "text_wrap": True}
        )
        text_cell = workbook.add_format({"font_color": "#263238", "valign": "top"})
        wrapped = workbook.add_format({"font_color": "#263238", "valign": "top", "text_wrap": True})
        numeric = workbook.add_format({"num_format": "0.0000", "align": "right"})
        probability = workbook.add_format({"num_format": "0.0000E+00", "align": "right"})
        count = workbook.add_format({"num_format": "#,##0", "align": "right"})
        id_format = workbook.add_format({"font_name": "Consolas", "font_size": 9})

        dashboard = workbook.add_worksheet(REPORT_SHEETS[0])
        dashboard.hide_gridlines(2)
        dashboard.set_tab_color("#1F4E78")
        dashboard.set_row(0, 30)
        dashboard.merge_range("A1:J1", "LOCAL ANALYTICS COPILOT — ANALYST REPORT", title)
        dashboard.merge_range(
            "A2:J2",
            "Deterministic association screening. Cards are ordered by adjusted p-value for "
            "screening only; they are not KPIs, causal effects, predictions or a cross-method "
            "business-importance ranking.",
            subtitle,
        )
        dashboard.set_row(1, 42)
        _write_text(dashboard, 3, 0, "Target column", section)
        _write_text(dashboard, 3, 1, payload["target_semantics"]["column"], text_cell)
        _write_text(dashboard, 3, 3, "Statistical role", section)
        _write_text(dashboard, 3, 4, payload["target_semantics"]["statistical_role"], text_cell)
        _write_text(dashboard, 3, 6, "Business meaning", section)
        _write_text(dashboard, 3, 7, "Unverified", note)
        dashboard.merge_range("A5:J5", payload["kpi_selection"]["reason"], note)
        dashboard_headers = [
            "Priority",
            "Predictor",
            "Method",
            "Effect measure",
            "Effect value",
            "Raw p-value",
            "Adjusted p-value",
            "Complete N",
            "Finding ID",
            "Deterministic source",
        ]
        for column, label in enumerate(dashboard_headers):
            _write_text(dashboard, 6, column, label, header)
        for index, card in enumerate(payload["dashboard"]["cards"], start=1):
            row = 6 + index
            analysis = analyses_by_effect[card["finding_id"]]
            ids = analysis["finding_ids"]
            dashboard.write_number(row, 0, index, count)
            _write_text(dashboard, row, 1, analysis["predictor"], text_cell)
            _write_text(dashboard, row, 2, analysis["method"], text_cell)
            _write_text(dashboard, row, 3, analysis["effect_name"], text_cell)
            formula_specs = [
                (4, ids["effect"], "D", numeric),
                (5, ids["p_value"], "D", probability),
                (6, ids["adjusted_p_value"], "D", probability),
                (7, ids["n"], "D", count),
                (8, ids["effect"], "A", id_format),
                (9, ids["effect"], "F", wrapped),
            ]
            for column, finding_id, evidence_column, cell_format in formula_specs:
                finding = finding_index[finding_id]
                cached_value = (
                    finding["value"]
                    if evidence_column == "D"
                    else finding_id
                    if evidence_column == "A"
                    else finding["source"]
                )
                dashboard.write_formula(
                    row,
                    column,
                    f"='Evidence'!${evidence_column}${evidence_rows[finding_id]}",
                    cell_format,
                    cached_value,
                )
        dashboard.freeze_panes(7, 0)
        dashboard.autofilter(6, 0, 6 + len(payload["dashboard"]["cards"]), 9)
        dashboard.set_column("A:A", 18)
        dashboard.set_column("B:C", 23)
        dashboard.set_column("D:D", 20)
        dashboard.set_column("E:H", 16)
        dashboard.set_column("I:I", 52)
        dashboard.set_column("J:J", 42)
        dashboard.repeat_rows(0, 6)
        dashboard.set_landscape()
        dashboard.fit_to_pages(1, 1)

        associations = workbook.add_worksheet(REPORT_SHEETS[1])
        associations.hide_gridlines(2)
        association_headers = [
            "Analysis ID",
            "Target",
            "Predictor",
            "Predictor kind",
            "Method",
            "Effect measure",
            "Effect value",
            "Raw p-value",
            "Adjusted p-value",
            "Complete N",
            "Assumption status",
            "Warning",
        ]
        for column, label in enumerate(association_headers):
            _write_text(associations, 0, column, label, header)
        for row, analysis in enumerate(payload["analyses"], start=1):
            ids = analysis["finding_ids"]
            values = [
                analysis["analysis_id"],
                analysis["target"],
                analysis["predictor"],
                analysis["predictor_kind"],
                analysis["method"],
                analysis["effect_name"],
            ]
            for column, value in enumerate(values):
                _write_text(associations, row, column, value, text_cell)
            associations.write_number(row, 6, finding_index[ids["effect"]]["value"], numeric)
            associations.write_number(row, 7, finding_index[ids["p_value"]]["value"], probability)
            associations.write_number(
                row, 8, finding_index[ids["adjusted_p_value"]]["value"], probability
            )
            associations.write_number(row, 9, finding_index[ids["n"]]["value"], count)
            _write_text(associations, row, 10, analysis["assumption_status"], text_cell)
            _write_text(associations, row, 11, analysis.get("warning", ""), wrapped)
        associations.autofilter(0, 0, len(payload["analyses"]), len(association_headers) - 1)
        associations.freeze_panes(1, 3)
        associations.set_column("A:A", 48)
        associations.set_column("B:F", 22)
        associations.set_column("G:J", 16)
        associations.set_column("K:K", 18)
        associations.set_column("L:L", 48)

        evidence = workbook.add_worksheet(REPORT_SHEETS[2])
        evidence.hide_gridlines(2)
        for column, label in enumerate(EVIDENCE_HEADERS):
            _write_text(evidence, 0, column, label, header)
        for row, finding in enumerate(findings, start=1):
            _write_text(evidence, row, 0, finding["finding_id"], id_format)
            _write_text(evidence, row, 1, finding["label"], text_cell)
            _write_text(evidence, row, 2, finding["kind"], text_cell)
            value_format = (
                probability
                if finding["unit"] in {"p_value", "adjusted_p_value"}
                else count
                if finding["unit"] in {"observations", "rows", "columns", "cells"}
                else numeric
            )
            evidence.write_number(row, 3, finding["value"], value_format)
            _write_text(evidence, row, 4, finding["unit"], text_cell)
            _write_text(evidence, row, 5, finding["source"], wrapped)
            _write_text(
                evidence,
                row,
                6,
                json.dumps(finding.get("dimension", {}), ensure_ascii=False, sort_keys=True),
                wrapped,
            )
            if finding.get("warning"):
                _write_text(evidence, row, 7, finding["warning"], wrapped)
            evidence.set_row(row, 36)
        evidence.autofilter(0, 0, len(findings), len(EVIDENCE_HEADERS) - 1)
        evidence.freeze_panes(1, 1)
        evidence.set_column("A:A", 58)
        evidence.set_column("B:C", 28)
        evidence.set_column("D:E", 18)
        evidence.set_column("F:H", 46)

        methodology = workbook.add_worksheet(REPORT_SHEETS[3])
        methodology.hide_gridlines(2)
        _write_text(methodology, 0, 0, "Methodology item", header)
        _write_text(methodology, 0, 1, "Value", header)
        methodology_rows = [
            ("Report schema", REPORT_SCHEMA_VERSION),
            ("Analyst schema", payload["schema_version"]),
            ("Target selection", "Explicit user selection"),
            ("Target business meaning", "Unverified"),
            ("Multiple-testing method", payload["multiple_testing"]["method"]),
            ("Multiple-testing family", payload["multiple_testing"]["family"]),
            ("Dashboard ranking", payload["dashboard"]["ranking_basis"]),
            ("Dashboard warning", payload["dashboard"]["warning"]),
            ("KPI status", payload["kpi_selection"]["status"]),
            ("KPI reason", payload["kpi_selection"]["reason"]),
            ("Deterministic verification", payload["verification"]["status"]),
            ("Interpretation status", payload["interpretation"]["status"]),
        ]
        interpretation = payload["interpretation"]
        if interpretation.get("status") == "completed":
            methodology_rows.extend(
                [
                    ("Interpretation model", interpretation.get("model", "")),
                    (
                        "Interpretation evidence IDs",
                        ", ".join(
                            interpreted for interpreted in interpretation["evidence_finding_ids"]
                        ),
                    ),
                    ("Verified interpretation", interpretation.get("text", "")),
                ]
            )
        for row, (key, value) in enumerate(methodology_rows, start=1):
            _write_text(methodology, row, 0, key, text_cell)
            _write_text(methodology, row, 1, value, wrapped)
            if len(_text(value)) > 100:
                methodology.set_row(row, 48 if key != "Verified interpretation" else 84)
        methodology.freeze_panes(1, 0)
        methodology.set_column("A:A", 32)
        methodology.set_column("B:B", 110)


def _same_number(actual: Any, expected: Any) -> bool:
    return (
        isinstance(actual, (int, float))
        and not isinstance(actual, bool)
        and math.isclose(float(actual), float(expected), rel_tol=1e-9, abs_tol=1e-12)
    )


def validate_analyst_excel_report(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    manifest_sha256 = analyst_manifest_sha256(payload)
    errors: list[dict[str, str]] = []
    try:
        formula_book = load_workbook(path, data_only=False, keep_links=False)
        value_book = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        return {
            "status": "failed",
            "scope": "analyst_xlsx_structure_formula_evidence_and_error_scan",
            "errors": [{"code": "workbook_open_failed", "message": str(exc)[:500]}],
            "sheets": [],
            "finding_count": 0,
            "dashboard_card_count": 0,
            "formula_count": 0,
            "error_cells": [],
            "external_links": [],
            "manifest_sha256": manifest_sha256,
        }

    if formula_book.sheetnames != REPORT_SHEETS:
        errors.append(
            {
                "code": "invalid_sheet_contract",
                "message": f"Expected {REPORT_SHEETS}, found {formula_book.sheetnames}",
            }
        )
    keywords = str(formula_book.properties.keywords or "")
    if f"lac-manifest-sha256={manifest_sha256}" not in keywords:
        errors.append(
            {
                "code": "manifest_digest_mismatch",
                "message": "Workbook metadata is not bound to the verified Analyst manifest",
            }
        )
    finding_index = _finding_index(payload)
    evidence_rows: dict[str, int] = {}
    evidence_max_row = 0
    if "Evidence" in formula_book.sheetnames:
        sheet = formula_book["Evidence"]
        evidence_max_row = sheet.max_row
        headers = [sheet.cell(1, column).value for column in range(1, len(EVIDENCE_HEADERS) + 1)]
        if headers != EVIDENCE_HEADERS:
            errors.append({"code": "invalid_evidence_headers", "message": str(headers)})
        for row in range(2, sheet.max_row + 1):
            finding_id = sheet.cell(row, 1).value
            if not isinstance(finding_id, str):
                continue
            if finding_id in evidence_rows:
                errors.append({"code": "duplicate_evidence_id", "message": finding_id})
                continue
            evidence_rows[finding_id] = row
            expected = finding_index.get(finding_id)
            if expected is None:
                errors.append({"code": "unknown_evidence_id", "message": finding_id})
                continue
            if not _same_number(sheet.cell(row, 4).value, expected["value"]):
                errors.append({"code": "evidence_value_mismatch", "message": finding_id})
            if sheet.cell(row, 5).value != expected["unit"]:
                errors.append({"code": "evidence_unit_mismatch", "message": finding_id})
            if sheet.cell(row, 6).value != expected["source"]:
                errors.append({"code": "evidence_source_mismatch", "message": finding_id})
        missing = sorted(set(finding_index) - set(evidence_rows))
        if missing:
            errors.append({"code": "missing_evidence_ids", "message": ", ".join(missing[:20])})

    formula_count = 0
    error_cells: list[str] = []
    external_links: list[str] = []
    for sheet in formula_book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in _FORMULA_ERRORS:
                    error_cells.append(f"{sheet.title}!{cell.coordinate}")
                if cell.data_type != "f":
                    continue
                formula_count += 1
                formula = str(value)
                if "[" in formula or "://" in formula:
                    external_links.append(f"{sheet.title}!{cell.coordinate}")
                match = _DIRECT_EVIDENCE_FORMULA.fullmatch(formula)
                if sheet.title != "Executive Dashboard" or not match:
                    errors.append(
                        {
                            "code": "invalid_formula_reference",
                            "message": f"{sheet.title}!{cell.coordinate}: {formula}",
                        }
                    )
                    continue
                referenced_row = int(match.group(2))
                if referenced_row < 2 or referenced_row > evidence_max_row:
                    errors.append(
                        {
                            "code": "formula_reference_out_of_range",
                            "message": f"{sheet.title}!{cell.coordinate}: {formula}",
                        }
                    )

    cards = payload.get("dashboard", {}).get("cards", [])
    analyses_by_effect = {
        analysis["finding_ids"]["effect"]: analysis for analysis in payload.get("analyses", [])
    }
    if "Executive Dashboard" in formula_book.sheetnames:
        formula_sheet = formula_book["Executive Dashboard"]
        value_sheet = value_book["Executive Dashboard"]
        for index, card in enumerate(cards, start=1):
            excel_row = 7 + index
            analysis = analyses_by_effect.get(card["finding_id"])
            if analysis is None:
                errors.append({"code": "dashboard_analysis_missing", "message": card["finding_id"]})
                continue
            ids = analysis["finding_ids"]
            expected_formulas = {
                "E": (ids["effect"], "D"),
                "F": (ids["p_value"], "D"),
                "G": (ids["adjusted_p_value"], "D"),
                "H": (ids["n"], "D"),
                "I": (ids["effect"], "A"),
                "J": (ids["effect"], "F"),
            }
            for column, (finding_id, evidence_column) in expected_formulas.items():
                expected_formula = (
                    f"='Evidence'!${evidence_column}${evidence_rows.get(finding_id, -1)}"
                )
                if formula_sheet[f"{column}{excel_row}"].value != expected_formula:
                    errors.append(
                        {
                            "code": "dashboard_formula_mismatch",
                            "message": f"{column}{excel_row}",
                        }
                    )
            numeric_cells = {
                "E": finding_index[ids["effect"]]["value"],
                "F": finding_index[ids["p_value"]]["value"],
                "G": finding_index[ids["adjusted_p_value"]]["value"],
                "H": finding_index[ids["n"]]["value"],
            }
            for column, expected in numeric_cells.items():
                if not _same_number(value_sheet[f"{column}{excel_row}"].value, expected):
                    errors.append(
                        {
                            "code": "dashboard_cached_value_mismatch",
                            "message": f"{column}{excel_row}",
                        }
                    )

    expected_formula_count = len(cards) * 6
    if formula_count != expected_formula_count:
        errors.append(
            {
                "code": "formula_count_mismatch",
                "message": f"Expected {expected_formula_count}, found {formula_count}",
            }
        )
    if error_cells:
        errors.append({"code": "workbook_error_cells", "message": ", ".join(error_cells[:20])})
    if external_links or getattr(formula_book, "_external_links", []):
        errors.append(
            {"code": "external_links_detected", "message": ", ".join(external_links[:20])}
        )
    formula_book.close()
    value_book.close()
    return {
        "status": "passed" if not errors else "failed",
        "scope": "analyst_xlsx_structure_formula_evidence_and_error_scan",
        "errors": errors,
        "sheets": REPORT_SHEETS,
        "finding_count": len(evidence_rows),
        "dashboard_card_count": len(cards),
        "formula_count": formula_count,
        "error_cells": error_cells,
        "external_links": external_links,
        "manifest_sha256": manifest_sha256,
    }


def create_analyst_excel_report(
    payload: dict[str, Any], output_name: str = "analyst_report.xlsx"
) -> dict[str, Any]:
    if (
        payload.get("schema_version") != "analyst.v1"
        or payload.get("verification", {}).get("status") != "passed"
    ):
        raise ValueError("Only verifier-passed analyst.v1 payloads can produce a report")
    output = safe_output_path(output_name, ".xlsx")
    _write_report_workbook(output, payload)
    verification = validate_analyst_excel_report(output, payload)
    if verification["status"] != "passed":
        output.unlink(missing_ok=True)
        raise RuntimeError(
            f"Analyst Excel report verification failed: {verification['errors'][:3]}"
        )
    settings = get_settings()
    relative_output = str(output.resolve().relative_to(settings.workspace.resolve()))
    audit(
        settings.logs_dir,
        "analyst_excel_report",
        output=relative_output,
        target=payload["target_semantics"]["column"],
        findings=verification["finding_count"],
    )
    return {
        "status": "created",
        "schema_version": REPORT_SCHEMA_VERSION,
        "output": relative_output,
        "filename": output.name,
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "verification": verification,
    }
