import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from lacopilot.analyst_interpretation import (
    analyst_digest,
    build_analyst_interpretation_messages,
    verify_analyst_interpretation,
)
from lacopilot.analyst_pipeline import run_analyst_pipeline, verify_analyst_payload
from lacopilot.analyst_report import (
    REPORT_SHEETS,
    create_analyst_excel_report,
    validate_analyst_excel_report,
)
from lacopilot.app import app
from lacopilot.config import get_settings
from lacopilot.regression_fixture import write_credit_risk_regression_fixture


def configure(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("LAC_WORKSPACE", str(tmp_path / "workspace"))
    monkeypatch.setenv("LAC_CONFIG_DIR", str(tmp_path / "config"))
    monkeypatch.delenv("LAC_API_TOKEN", raising=False)
    get_settings.cache_clear()
    return get_settings()


def _finding_values(payload):
    return {finding["finding_id"]: finding["value"] for finding in payload["findings"]}


def test_binary_target_analyst_contract_matches_csv_and_xlsx(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    paths = write_credit_risk_regression_fixture(settings.incoming_dir)
    predictors = ["utilization_rate", "customer_segment", "legal_status"]

    csv_payload = run_analyst_pipeline(
        f"incoming/{paths['csv'].name}",
        "default_next_30d",
        predictor_columns=predictors,
    )
    xlsx_payload = run_analyst_pipeline(
        f"incoming/{paths['xlsx'].name}",
        "default_next_30d",
        predictor_columns=predictors,
    )

    for payload in (csv_payload, xlsx_payload):
        assert payload["status"] == "completed"
        assert payload["mode"] == "analyst"
        assert payload["target_semantics"] == {
            "column": "default_next_30d",
            "statistical_role": "binary",
            "selection_source": "explicit_request",
            "business_meaning_status": "unverified",
            "business_meaning": None,
        }
        assert payload["kpi_selection"]["status"] == "requires_approved_definition"
        assert payload["kpi_selection"]["selected"] == []
        assert payload["verification"]["status"] == "passed"
        analyses = {analysis["predictor"]: analysis for analysis in payload["analyses"]}
        assert analyses["utilization_rate"]["method"] == "mann_whitney_u"
        assert analyses["customer_segment"]["method"] == "chi_square"
        assert analyses["legal_status"]["method"] == "chi_square"
        assert analyses["utilization_rate"]["finding_ids"]["effect"] == (
            "analyst.target.21.association.column.10.effect"
        )
        findings = {finding["finding_id"]: finding for finding in payload["findings"]}
        assert all(finding["source"] for finding in findings.values())
        assert all(card["finding_id"] in findings for card in payload["dashboard"]["cards"])
        assert payload["dashboard"]["evidence_policy"] == ("all_numeric_cards_bound_to_finding_id")
        assert payload["dashboard"]["ranking_basis"] == ("adjusted_p_value_then_finding_id")
        first_effect_id = payload["dashboard"]["cards"][0]["finding_id"]
        first_adjusted_id = f"{first_effect_id[:-7]}.adjusted_p_value"
        adjusted_values = [
            finding["value"]
            for finding in payload["findings"]
            if finding["unit"] == "adjusted_p_value"
        ]
        assert findings[first_adjusted_id]["value"] == min(adjusted_values)
        assert "probability" not in str(payload["target_semantics"]).lower()

    csv_values = _finding_values(csv_payload)
    xlsx_values = _finding_values(xlsx_payload)
    assert csv_values.keys() == xlsx_values.keys()
    for finding_id, csv_value in csv_values.items():
        assert xlsx_values[finding_id] == pytest.approx(csv_value)


def test_nonbinary_target_requires_explicit_statistical_kind(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "continuous.csv"
    pd.DataFrame({"target": range(1, 21), "predictor": range(21, 41)}).to_csv(path, index=False)

    with pytest.raises(ValueError, match="target_kind"):
        run_analyst_pipeline("incoming/continuous.csv", "target")

    payload = run_analyst_pipeline(
        "incoming/continuous.csv",
        "target",
        target_kind="continuous",
        predictor_columns=["predictor"],
    )
    assert payload["analyses"][0]["method"] == "spearman"
    effect_id = payload["analyses"][0]["finding_ids"]["effect"]
    assert _finding_values(payload)[effect_id] == pytest.approx(1.0)


def test_default_predictor_selection_excludes_identifiers_and_target(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame(
        {
            "customer_id": [f"C-{index:03d}" for index in range(20)],
            "target": [0] * 10 + [1] * 10,
            "value": list(range(20)),
            "segment": ["A", "B"] * 10,
        }
    ).to_csv(settings.incoming_dir / "selection.csv", index=False)

    payload = run_analyst_pipeline("incoming/selection.csv", "target")

    assert payload["predictor_selection"]["source"] == "deterministic_role_filter"
    assert payload["predictor_selection"]["included"] == ["value", "segment"]
    assert {analysis["predictor"] for analysis in payload["analyses"]} == {
        "value",
        "segment",
    }


def test_analyst_verifier_rejects_unbound_card_and_business_semantics(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "binary.csv"
    pd.DataFrame(
        {
            "target": [0] * 10 + [1] * 10,
            "predictor": list(range(20)),
        }
    ).to_csv(path, index=False)
    payload = run_analyst_pipeline("incoming/binary.csv", "target", predictor_columns=["predictor"])

    payload["dashboard"]["cards"][0]["value"] = 999
    payload["findings"][0]["value"] = "not-a-number"
    payload["target_semantics"]["business_meaning_status"] = "inferred"
    verification = verify_analyst_payload(payload)

    assert verification["status"] == "failed"
    assert {error["code"] for error in verification["errors"]} >= {
        "unbound_dashboard_card",
        "invalid_numeric_value",
        "unsupported_business_semantics",
    }


def test_analyst_api_returns_typed_invalid_request_error(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"target": [0, 1, 0, 1], "value": [1, 2, 3, 4]}).to_csv(
        settings.incoming_dir / "api.csv", index=False
    )
    client = TestClient(app)

    response = client.post(
        "/api/v1/analysis/analyst",
        json={"file_path": "incoming/api.csv", "target_column": "missing"},
    )

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "invalid_analysis_request"


def _interpretation_payload(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame(
        {
            "target": [0] * 12 + [1] * 12,
            "predictor": list(range(24)),
            "unused_raw_value": [f"raw-{index}" for index in range(24)],
        }
    ).to_csv(settings.incoming_dir / "interpretation.csv", index=False)
    return run_analyst_pipeline(
        "incoming/interpretation.csv",
        "target",
        predictor_columns=["predictor"],
    )


def test_analyst_interpretation_digest_and_verifier_accept_bound_claims(tmp_path, monkeypatch):
    payload = _interpretation_payload(tmp_path, monkeypatch)
    analysis = payload["analyses"][0]
    finding_ids = analysis["finding_ids"]
    findings = {finding["finding_id"]: finding for finding in payload["findings"]}
    effect = findings[finding_ids["effect"]]["value"]
    adjusted = findings[finding_ids["adjusted_p_value"]]["value"]
    sample_size = findings[finding_ids["n"]]["value"]
    text = (
        f"Etki ölçüsü {effect} [{finding_ids['effect']}], düzeltilmiş p-değeri "
        f"{adjusted} [{finding_ids['adjusted_p_value']}] ve gözlem sayısı "
        f"{sample_size} [{finding_ids['n']}]. "
        "İş anlamı belirlenemiyor ve bu ilişki nedensellik göstermez."
    )

    digest = analyst_digest(payload)
    verification = verify_analyst_interpretation(text, payload)

    assert verification["status"] == "passed"
    assert set(verification["cited_finding_ids"]) == {
        finding_ids["effect"],
        finding_ids["adjusted_p_value"],
        finding_ids["n"],
    }
    assert "unused_raw_value" not in str(digest)
    assert {finding["finding_id"] for finding in digest["evidence"]} == set(finding_ids.values())
    messages = build_analyst_interpretation_messages(payload, "İlişkileri açıkla")
    assert "Never recalculate" in messages[0]["content"]
    assert "raw-0" not in messages[1]["content"]


def test_analyst_interpretation_verifier_rejects_invented_numbers_and_semantics(
    tmp_path, monkeypatch
):
    payload = _interpretation_payload(tmp_path, monkeypatch)
    effect_id = payload["analyses"][0]["finding_ids"]["effect"]
    text = (
        f"Predictor 98% ile en önemli sürücüdür ve temerrüt olasılığını tahmin eder "
        f"[{effect_id}]. Sonuç istatistiksel olarak anlamlıdır. Ek skor 777'dir."
    )

    verification = verify_analyst_interpretation(text, payload)
    semantic_codes = {violation["code"] for violation in verification["semantic_violations"]}

    assert verification["status"] == "needs_review"
    assert verification["numeric_evidence_mismatches"]
    assert verification["uncited_numeric_claims"]
    assert semantic_codes >= {
        "causal_claim",
        "business_importance_claim",
        "unsupported_significance_threshold",
        "unsupported_prediction_semantics",
    }


def test_analyst_pipeline_hides_rejected_local_model_text(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame(
        {
            "target": [0] * 12 + [1] * 12,
            "predictor": list(range(24)),
        }
    ).to_csv(settings.incoming_dir / "rejected.csv", index=False)

    class FakeClient:
        def __init__(self, **_kwargs):
            pass

        def chat(self, **_kwargs):
            return SimpleNamespace(
                message=SimpleNamespace(content="Bu değişken 99% ile en önemli risk sürücüsüdür.")
            )

    monkeypatch.setitem(sys.modules, "ollama", SimpleNamespace(Client=FakeClient))
    payload = run_analyst_pipeline(
        "incoming/rejected.csv",
        "target",
        predictor_columns=["predictor"],
        interpret=True,
    )

    assert payload["verification"]["status"] == "passed"
    assert payload["interpretation"]["status"] == "rejected"
    assert "text" not in payload["interpretation"]
    assert payload["interpretation"]["verification"]["status"] == "needs_review"


def test_analyst_excel_report_reopens_with_finding_bound_formulas(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame(
        {
            "target": [0] * 12 + [1] * 12,
            "predictor": list(range(24)),
            "raw_secret": [f"never-export-{index}" for index in range(24)],
        }
    ).to_csv(settings.incoming_dir / "report.csv", index=False)
    payload = run_analyst_pipeline("incoming/report.csv", "target", predictor_columns=["predictor"])

    result = create_analyst_excel_report(payload, "verified_analyst.xlsx")
    output = settings.workspace / result["output"]
    formula_book = load_workbook(output, data_only=False)
    value_book = load_workbook(output, data_only=True)

    assert result["verification"]["status"] == "passed"
    assert result["verification"]["error_cells"] == []
    assert result["verification"]["external_links"] == []
    assert formula_book.sheetnames == REPORT_SHEETS
    assert formula_book["Executive Dashboard"]["E8"].value == "='Evidence'!$D$2"
    effect_id = payload["analyses"][0]["finding_ids"]["effect"]
    effect = _finding_values(payload)[effect_id]
    assert value_book["Executive Dashboard"]["E8"].value == pytest.approx(effect)
    all_text = " ".join(
        str(cell.value)
        for sheet in formula_book.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "never-export" not in all_text

    formula_book["Executive Dashboard"]["E8"] = "=1/0"
    formula_book.save(output)
    tampered = validate_analyst_excel_report(output, payload)
    assert tampered["status"] == "failed"
    assert "invalid_formula_reference" in {error["code"] for error in tampered["errors"]}

    formula_book.remove(formula_book["Evidence"])
    formula_book.save(output)
    missing_evidence = validate_analyst_excel_report(output, payload)
    assert missing_evidence["status"] == "failed"
    assert "invalid_sheet_contract" in {error["code"] for error in missing_evidence["errors"]}


def test_analyst_excel_report_treats_formula_like_column_as_text(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"target": [0] * 12 + [1] * 12, "=2+2": list(range(24))}).to_csv(
        settings.incoming_dir / "formula-name.csv", index=False
    )
    payload = run_analyst_pipeline(
        "incoming/formula-name.csv", "target", predictor_columns=["=2+2"]
    )

    result = create_analyst_excel_report(payload, "safe_formula_name.xlsx")
    workbook = load_workbook(settings.workspace / result["output"], data_only=False)

    assert workbook["Associations"]["C2"].value == "=2+2"
    assert workbook["Associations"]["C2"].data_type == "s"


def test_analyst_report_api_returns_only_a_verified_xlsx(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"target": [0] * 12 + [1] * 12, "predictor": list(range(24))}).to_csv(
        settings.incoming_dir / "api-report.csv", index=False
    )
    client = TestClient(app)

    response = client.post(
        "/api/v1/analysis/analyst/report",
        json={
            "file_path": "incoming/api-report.csv",
            "target_column": "target",
            "predictor_columns": ["predictor"],
            "interpret": False,
            "output_name": "../api_verified.xlsx",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["x-lac-report-verification"] == "passed"
    assert "api_verified.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == REPORT_SHEETS
    assert (settings.outputs_dir / "api_verified.xlsx").is_file()
    assert not (tmp_path / "api_verified.xlsx").exists()
