import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from lacopilot.app import app
from lacopilot.config import get_settings
from lacopilot.ingestion import (
    IngestionError,
    detect_csv_dialect,
    inspect_workbook,
    read_table,
    validate_excel_archive,
)
from lacopilot.quick_analysis import (
    build_interpretation_messages,
    build_quick_dashboard,
    interpret_profile,
    profile_digest,
    verify_interpretation,
)
from lacopilot.regression_fixture import write_credit_risk_regression_fixture
from lacopilot.tools.data_tools import profile_dataset


def configure(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("LAC_WORKSPACE", str(tmp_path / "workspace"))
    monkeypatch.setenv("LAC_CONFIG_DIR", str(tmp_path / "config"))
    monkeypatch.delenv("LAC_API_TOKEN", raising=False)
    get_settings.cache_clear()
    return get_settings()


def test_csv_detection_handles_turkish_encoding_and_decimal_comma(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "tahsilat.csv"
    path.write_bytes("müşteri;tutar;şehir\nA;1.234,50;İstanbul\nB;950,25;İzmir\n".encode("cp1254"))

    dialect = detect_csv_dialect(path)
    loaded = read_table(path)

    assert dialect.encoding == "cp1254"
    assert dialect.delimiter == ";"
    assert dialect.decimal_separator == ","
    assert dialect.thousands_separator == "."
    assert loaded.dataframe.shape == (2, 3)
    assert loaded.dataframe["tutar"].tolist() == [1234.5, 950.25]


def test_csv_malformed_rows_fail_instead_of_silent_drop(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "broken.csv"
    path.write_text('a,b\n1,"unterminated\n2,3\n', encoding="utf-8")

    with pytest.raises(IngestionError) as caught:
        read_table(path)

    assert caught.value.code in {"malformed_csv", "csv_shape_mismatch"}


def test_csv_headers_are_nonempty_and_unique_without_name_collisions(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "headers.csv"
    path.write_text("a,a_2,a,,\n1,2,3,4,5\n", encoding="utf-8")

    loaded = read_table(path)

    assert list(loaded.dataframe.columns) == ["a", "a_2", "a_3", "column_4", "column_5"]
    assert len(set(loaded.dataframe.columns)) == 5


def test_csv_missing_fields_are_reported_with_row_evidence(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "short-row.csv"
    path.write_text("a,b,c\n1,2,3\n4,5\n", encoding="utf-8")

    with pytest.raises(IngestionError) as caught:
        read_table(path)

    assert caught.value.code == "inconsistent_csv_row"
    assert caught.value.details["logical_row"] == 3
    assert caught.value.details["expected_columns"] == 3
    assert caught.value.details["actual_columns"] == 2


def test_excel_discovers_sheets_and_header_rows(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    path = settings.incoming_dir / "workbook.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Risk Data"
    first.append(["Aylık Risk Raporu"])
    first.append([])
    first.append(["customer_id", "monthly_income_try", "default_next_30d"])
    first.append(["C1", 25_000, 0])
    first.append(["C2", 18_000, 1])
    notes = workbook.create_sheet("Notes")
    notes.append(["key", "value"])
    notes.append(["owner", "Risk"])
    workbook.save(path)

    manifest = inspect_workbook(path)
    loaded = read_table(path, "Risk Data")

    assert [sheet["name"] for sheet in manifest["sheets"]] == ["Risk Data", "Notes"]
    assert manifest["sheets"][0]["header_row"] == 3
    assert loaded.dataframe.shape == (2, 3)
    assert loaded.metadata["excel"]["selected_sheet"] == "Risk Data"

    with pytest.raises(IngestionError) as caught:
        validate_excel_archive(path, limit_bytes=1)
    assert caught.value.code == "excel_decompression_limit"


@pytest.mark.parametrize("kind", ["csv", "xlsx"])
def test_credit_risk_regression_fixture_contract(tmp_path, monkeypatch, kind):
    settings = configure(tmp_path, monkeypatch)
    paths = write_credit_risk_regression_fixture(settings.incoming_dir)
    profile = profile_dataset(f"incoming/{paths[kind].name}")

    assert profile["rows"] == 1508
    assert profile["columns"] == 22
    assert profile["total_missing_cells"] == 52
    assert profile["missing_cell_pct"] == pytest.approx(0.1567)
    assert profile["missing_count"]["monthly_income_try"] == 24
    assert profile["missing_count"]["payment_ratio_3m"] == 12
    assert profile["missing_count"]["employment_years"] == 16
    assert profile["duplicate_rows"] == 8
    assert profile["duplicate_rows_including_originals"] == 16
    assert profile["ingestion"]["parser"] == "lac-deterministic-data-bridge"
    dashboard = build_quick_dashboard(profile)
    cards = {card["finding_id"]: card for card in dashboard["cards"]}
    assert cards["profile.shape.rows"]["value"] == 1508
    assert cards["profile.shape.columns"]["value"] == 22
    assert cards["profile.quality.missing_cells"]["value"] == 52
    assert cards["profile.quality.exact_duplicate_copies"]["value"] == 8
    findings = {finding["finding_id"]: finding for finding in profile["findings"]}
    assert findings["profile.quality.missing_cell_rate"]["value"] == pytest.approx(0.1567)
    assert findings["profile.quality.duplicate_group_rows"]["value"] == 16
    assert dashboard["missing_by_column"] == [
        {
            "finding_id": "profile.quality.missing.column.6",
            "pct_finding_id": "profile.quality.missing_pct.column.6",
            "column": "monthly_income_try",
            "count": 24,
            "pct": 1.59,
        },
        {
            "finding_id": "profile.quality.missing.column.7",
            "pct_finding_id": "profile.quality.missing_pct.column.7",
            "column": "employment_years",
            "count": 16,
            "pct": 1.06,
        },
        {
            "finding_id": "profile.quality.missing.column.12",
            "pct_finding_id": "profile.quality.missing_pct.column.12",
            "column": "payment_ratio_3m",
            "count": 12,
            "pct": 0.8,
        },
    ]


def test_upload_api_profiles_csv_and_returns_stable_findings(tmp_path, monkeypatch):
    configure(tmp_path, monkeypatch)
    client = TestClient(app)
    content = b"customer_id,amount\nC1,100\nC2,200\nC2,200\n"

    response = client.post(
        "/api/v1/datasets/upload",
        files={"file": ("collections.csv", content, "text/csv")},
        data={"run_profile": "true", "interpret": "false"},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["status"] == "profiled"
    assert payload["profile"]["rows"] == 3
    assert payload["profile"]["duplicate_rows"] == 1
    assert {finding["finding_id"] for finding in payload["profile"]["findings"]} >= {
        "profile.shape.rows",
        "profile.shape.columns",
        "profile.quality.missing_cells",
        "profile.quality.exact_duplicate_copies",
    }
    assert [card["finding_id"] for card in payload["dashboard"]["cards"]] == [
        "profile.shape.rows",
        "profile.shape.columns",
        "profile.quality.missing_cells",
        "profile.quality.exact_duplicate_copies",
        "profile.quality.score_heuristic",
    ]
    assert payload["dashboard"]["evidence_policy"] == ("all_numeric_cards_bound_to_finding_id")


def test_upload_api_never_silently_accepts_wrong_excel_signature(tmp_path, monkeypatch):
    configure(tmp_path, monkeypatch)
    client = TestClient(app)

    response = client.post(
        "/api/v1/datasets/upload",
        files={"file": ("fake.xlsx", b"not an excel workbook", "application/octet-stream")},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == {
        "code": "signature_mismatch",
        "message": "Dosya uzantısı Excel olsa da içerik geçerli bir XLSX/XLSM paketi değil.",
        "hint": "Dosyayı Excel'de yeniden .xlsx olarak kaydedin.",
        "details": {},
    }
    assert not list(get_settings().incoming_dir.glob("fake*.xlsx"))


def test_upload_api_requests_sheet_selection_for_multi_sheet_excel(tmp_path, monkeypatch):
    configure(tmp_path, monkeypatch)
    workbook = Workbook()
    first = workbook.active
    first.title = "Ocak"
    first.append(["id", "amount"])
    first.append([1, 100])
    second = workbook.create_sheet("Şubat")
    second.append(["id", "amount"])
    second.append([2, 200])
    content = BytesIO()
    workbook.save(content)

    response = TestClient(app).post(
        "/api/v1/datasets/upload",
        files={
            "file": (
                "collections.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["status"] == "sheet_selection_required"
    assert [sheet["name"] for sheet in payload["sheet_options"]] == ["Ocak", "Şubat"]


def test_quick_interpretation_prompt_contains_facts_not_raw_rows(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"customer_id": ["C1", "C2"], "amount": [10, None]}).to_csv(
        settings.incoming_dir / "small.csv", index=False
    )
    profile = profile_dataset("incoming/small.csv")
    messages = build_interpretation_messages(profile, "Veri kalitesi nasıl?")
    digest = profile_digest(profile)

    evidence = {finding["finding_id"]: finding for finding in digest["evidence"]}
    assert evidence["profile.quality.missing_cells"]["value"] == 1
    assert evidence["profile.quality.missing_cell_rate"]["unit"] == "percent_of_all_cells"
    assert "profile.quality.missing_cells" in messages[1]["content"]
    assert "NEVER add column-level missing percentages" in messages[0]["content"]
    assert 'explicitly say "belirlenemiyor"' in messages[0]["content"]
    assert "C1" not in messages[1]["content"]
    assert "C2" not in messages[1]["content"]


def test_quick_dashboard_is_deterministic_and_contains_no_raw_rows(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame(
        {
            "customer_id": ["C1", "C2", "C3"],
            "amount": [10, None, None],
            "segment": ["A", "B", "A"],
        }
    ).to_csv(settings.incoming_dir / "small.csv", index=False)
    profile = profile_dataset("incoming/small.csv")

    dashboard = build_quick_dashboard(profile)

    assert dashboard["dashboard_version"] == 1
    assert dashboard["missing_by_column"] == [
        {
            "finding_id": "profile.quality.missing.column.1",
            "pct_finding_id": "profile.quality.missing_pct.column.1",
            "column": "amount",
            "count": 2,
            "pct": 66.67,
        }
    ]
    assert dashboard["role_counts"]
    assert "sample" not in dashboard
    assert "sample_rows" not in dashboard
    assert "C1" not in str(dashboard)


def test_quick_api_returns_dashboard_with_stable_finding_sources(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"amount": [10, 20, 20]}).to_csv(settings.incoming_dir / "small.csv", index=False)

    response = TestClient(app).post(
        "/api/v1/analysis/quick",
        json={"file_path": "incoming/small.csv", "interpret": False},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "completed"
    assert payload["interpretation"] == {"status": "skipped"}
    sources = {card["finding_id"]: card["source"] for card in payload["dashboard"]["cards"]}
    assert sources == {
        "profile.shape.rows": "deterministic_dataframe_shape",
        "profile.shape.columns": "deterministic_dataframe_shape",
        "profile.quality.missing_cells": "dataframe_isna_sum",
        "profile.quality.exact_duplicate_copies": "dataframe_duplicated_keep_first",
        "profile.quality.score_heuristic": "documented_screening_heuristic",
    }


def test_quick_interpretation_retries_models_without_think_option(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    pd.DataFrame({"amount": [10, 20]}).to_csv(settings.incoming_dir / "small.csv", index=False)
    profile = profile_dataset("incoming/small.csv")

    class FakeClient:
        calls = 0

        def __init__(self, **_kwargs):
            pass

        def chat(self, **kwargs):
            type(self).calls += 1
            if "think" in kwargs:
                raise ValueError("think option unsupported")
            return SimpleNamespace(
                message=SimpleNamespace(content="Satır sayısı 2 [profile.shape.rows].")
            )

    monkeypatch.setitem(sys.modules, "ollama", SimpleNamespace(Client=FakeClient))
    result = interpret_profile(profile)

    assert result["status"] == "completed"
    assert result["text"] == "Satır sayısı 2 [profile.shape.rows]."
    assert result["verification"]["status"] == "passed"
    assert result["evidence_finding_ids"] == ["profile.shape.rows"]
    assert verify_interpretation("Satır sayısı 2.", profile)["status"] == "needs_review"
    assert FakeClient.calls == 2


def test_interpretation_contract_rejects_duplicate_missing_and_binary_semantic_errors(
    tmp_path, monkeypatch
):
    settings = configure(tmp_path, monkeypatch)
    paths = write_credit_risk_regression_fixture(settings.incoming_dir)
    profile = profile_dataset(f"incoming/{paths['csv'].name}")
    digest = profile_digest(profile)

    binary_columns = {item["column"]: item for item in digest["binary_columns"]}
    assert binary_columns["default_next_30d"] == {
        "column": "default_next_30d",
        "classification": "binary_observed_values",
        "technical_role": "numeric",
        "business_meaning": "unknown_without_approved_metadata",
        "probability_interpretation_supported": False,
    }

    wrong_duplicate = (
        "Duplicate gruplarında 16 satır kaldırılmalı [profile.quality.duplicate_group_rows]."
    )
    duplicate_check = verify_interpretation(wrong_duplicate, profile)
    assert duplicate_check["status"] == "needs_review"
    assert {item["code"] for item in duplicate_check["semantic_violations"]} == {
        "duplicate_group_rows_used_as_removal_count"
    }

    wrong_missing = (
        "monthly_income_try için %1,59 [profile.quality.missing_pct.column.6] ve "
        "employment_years için %1,06 [profile.quality.missing_pct.column.7] olduğundan "
        "toplam eksik veri oranı %2,65'tir."
    )
    missing_check = verify_interpretation(wrong_missing, profile)
    assert missing_check["status"] == "needs_review"
    assert missing_check["numeric_evidence_mismatches"]
    assert {item["code"] for item in missing_check["semantic_violations"]} == {
        "column_missing_percentages_used_as_overall_rate"
    }

    wrong_binary = "default_next_30d, 30 gün içinde varsayılan olma olasılığı tahminidir."
    binary_check = verify_interpretation(wrong_binary, profile)
    assert binary_check["status"] == "needs_review"
    assert {item["code"] for item in binary_check["semantic_violations"]} == {
        "unsupported_binary_business_semantics"
    }
    unsupported_target = verify_interpretation(
        "default_next_30d ikili bir hedef sütunudur ama olasılık değildir.", profile
    )
    assert unsupported_target["status"] == "needs_review"
    assert unsupported_target["semantic_violations"][0]["code"] == (
        "unsupported_binary_business_semantics"
    )


def test_interpretation_contract_accepts_bound_facts_and_explicit_unknown_semantics(
    tmp_path, monkeypatch
):
    settings = configure(tmp_path, monkeypatch)
    paths = write_credit_risk_regression_fixture(settings.incoming_dir)
    profile = profile_dataset(f"incoming/{paths['csv'].name}")
    text = "\n".join(
        [
            "Fazladan duplicate kopya sayısı 8 [profile.quality.exact_duplicate_copies].",
            "Duplicate gruplarında orijinaller dahil 16 satır bulunur [profile.quality.duplicate_group_rows].",
            "Tüm hücrelerde eksik oranı %0,1567 [profile.quality.missing_cell_rate].",
            "default_next_30d ikili bir sütundur; iş anlamı profilden belirlenemiyor.",
        ]
    )

    check = verify_interpretation(text, profile)

    assert check["status"] == "passed"
    assert check["numeric_evidence_mismatches"] == []
    assert check["semantic_violations"] == []


def test_interpret_profile_hides_model_text_that_fails_semantic_verification(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    paths = write_credit_risk_regression_fixture(settings.incoming_dir)
    profile = profile_dataset(f"incoming/{paths['csv'].name}")

    class FakeClient:
        def __init__(self, **_kwargs):
            pass

        def chat(self, **_kwargs):
            return SimpleNamespace(
                message=SimpleNamespace(
                    content=(
                        "Duplicate gruplarında 16 satır kaldırılmalı "
                        "[profile.quality.duplicate_group_rows]."
                    )
                )
            )

    monkeypatch.setitem(sys.modules, "ollama", SimpleNamespace(Client=FakeClient))

    result = interpret_profile(profile)

    assert result["status"] == "rejected"
    assert "text" not in result
    assert result["verification"]["status"] == "needs_review"
    assert result["verification"]["semantic_violations"][0]["code"] == (
        "duplicate_group_rows_used_as_removal_count"
    )


def test_empty_dataset_profile_reports_zero_total_cells(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    (settings.incoming_dir / "headers-only.csv").write_text("a,b\n", encoding="utf-8")

    profile = profile_dataset("incoming/headers-only.csv")

    assert profile["rows"] == 0
    assert profile["columns"] == 2
    assert profile["total_cells"] == 0
    assert profile["total_missing_cells"] == 0
