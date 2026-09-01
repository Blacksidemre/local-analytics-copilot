from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from lacopilot.agent_report import (
    create_agent_excel_report,
    create_agent_html_report,
    create_agent_pdf_report,
    validate_agent_excel_report,
)
from lacopilot.config import get_settings


def configure(tmp_path, monkeypatch):
    monkeypatch.setenv("LAC_WORKSPACE", str(tmp_path / "workspace"))
    monkeypatch.setenv("LAC_CONFIG_DIR", str(tmp_path / "config"))
    get_settings.cache_clear()
    return get_settings()


def verified_history_run() -> dict:
    return {
        "run_id": "a" * 32,
        "dataset_id": "b" * 64,
        "dataset_ref": "incoming/fixture.csv",
        "source_name": "fixture.csv",
        "sheet_name": "0",
        "mode": "agent",
        "question": 'Bu veriyi özetle. <script src="https://example.invalid/x.js"></script>',
        "run_status": "completed",
        "verifier_status": "passed",
        "finding_count": 2,
        "created_at": "2026-08-31T12:00:00+00:00",
        "tools": ["profile_dataset"],
        "findings": [
            {
                "finding_id": "profile.shape.rows",
                "kind": "metric",
                "label": '=WEBSERVICE("https://example.invalid")',
                "value": 1508,
                "unit": "rows",
                "source": "deterministic_dataframe_shape",
            },
            {
                "finding_id": "profile.quality.exact_duplicate_copies",
                "kind": "metric",
                "label": "Fazladan exact duplicate kopyaları",
                "value": 8,
                "unit": "rows",
                "source": "deterministic_dataframe_duplicated_keep_first",
                "dimension": {"scope": "dataset"},
                "warning": "Original rows are not counted as removable copies.",
            },
        ],
    }


def output_path(settings, report: dict) -> Path:
    return settings.workspace / report["output"]


def test_agent_reports_share_one_verified_manifest_across_xlsx_html_pdf(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    run = verified_history_run()

    reports = [
        create_agent_excel_report(run),
        create_agent_html_report(run),
        create_agent_pdf_report(run),
    ]

    assert [report["schema_version"] for report in reports] == [
        "agent-report.v1",
        "agent-html-report.v1",
        "agent-pdf-report.v1",
    ]
    assert all(report["verification"]["status"] == "passed" for report in reports)
    assert all(report["verification"]["finding_count"] == 2 for report in reports)
    assert all(report["verification"]["dashboard_card_count"] == 0 for report in reports)
    assert len({report["verification"]["manifest_sha256"] for report in reports}) == 1
    assert all(output_path(settings, report).is_file() for report in reports)

    html = output_path(settings, reports[1]).read_text(encoding="utf-8")
    assert "<script" not in html
    assert "https://example.invalid" in html
    assert "&lt;script" in html

    workbook = load_workbook(output_path(settings, reports[0]), data_only=False, keep_links=False)
    assert workbook["Evidence"]["B2"].value.startswith("=WEBSERVICE")
    assert workbook["Evidence"]["B2"].data_type == "s"
    assert workbook._external_links == []
    workbook.close()


def test_agent_report_rejects_unverified_tampered_or_unsafe_manifest(tmp_path, monkeypatch):
    configure(tmp_path, monkeypatch)
    run = verified_history_run()

    unverified = run | {"verifier_status": "failed"}
    with pytest.raises(ValueError, match="verifier-passed"):
        create_agent_excel_report(unverified)

    mismatched = run | {"finding_count": 3}
    with pytest.raises(ValueError, match="finding_count"):
        create_agent_html_report(mismatched)

    duplicate = run | {"findings": [run["findings"][0], run["findings"][0]]}
    with pytest.raises(ValueError, match="unique"):
        create_agent_pdf_report(duplicate)

    with pytest.raises(ValueError, match="filename"):
        create_agent_excel_report(run, "../outside.xlsx")
    with pytest.raises(ValueError, match="filename"):
        create_agent_excel_report(run, "..\\outside.xlsx")


def test_agent_excel_validator_fails_closed_after_evidence_tampering(tmp_path, monkeypatch):
    settings = configure(tmp_path, monkeypatch)
    run = verified_history_run()
    report = create_agent_excel_report(run)
    path = output_path(settings, report)

    workbook = load_workbook(path)
    workbook["Evidence"]["D2"] = 999
    workbook.save(path)
    workbook.close()

    verification = validate_agent_excel_report(path, run)
    assert verification["status"] == "failed"
    assert {error["code"] for error in verification["errors"]} == {"evidence_value_mismatch"}
